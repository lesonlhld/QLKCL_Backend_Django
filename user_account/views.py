import os
import datetime, pytz
import locale
import json
import requests
import openpyxl, csv, codecs
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Avg, Sum, Q
from rest_framework import permissions
from rest_framework.decorators import action, permission_classes
from .validators.user import UserValidator
from .validators.home import HomeValidator
from .validators.destination_history import DestinationHistoryValidator
from .validators.quarantine_history import QuarantineHistoryValidator
from .models import CustomUser, Member, Manager, Staff, DestinationHistory, QuarantineHistory
from .serializers import (
    DestinationHistorySerializer, QuarantineHistorySerializer,
    CustomUserSerializer, MemberSerializer,
    FilterMemberSerializer, FilterNotMemberSerializer,
    MemberHomeSerializer, ManagerSerializer,
    StaffSerializer, FilterStaffSerializer, FilterManagerSerializer,
    BaseQuarantineHistorySerializer,
)
from .filters.member import MemberFilter
from .filters.user import UserFilter
from .filters.staff import StaffFilter
from .filters.manager import ManagerFilter
from .filters.destination_history import DestinationHistoryFilter
from .filters.quarantine_history import QuarantineHistoryFilter
from form.models import Test, VaccineDose, Pandemic, MedicalDeclaration, BackgroundDisease
from form.serializers import (
    BaseMedicalDeclarationSerializer,
    BaseTestSerializer,
)
from form.filters.test import TestFilter
from form.serializers import (
    BaseVaccineDoseSerializer,
)
from role.models import Role
from address.models import City, District, Ward
from quarantine_ward.models import QuarantineRoom
from quarantine_ward.serializers import (
    QuarantineRoomSerializer, QuarantineFloorSerializer,
    QuarantineBuildingSerializer, BaseQuarantineWardSerializer,
)
from notification.views import create_and_send_noti_to_list_user
from utils import exceptions, messages
from utils.enums import (
    CustomUserStatus, HealthStatus, TestStatus,
    MemberQuarantinedStatus, MemberLabel,
    QuarantineHistoryStatus, QuarantineHistoryEndType,
    Professional, TestResult, Gender,
)
from utils.views import AbstractView, paginate_data, query_debugger
from utils.tools import custom_user_code_generator, LabelTool, split_input_list

# Create your views here.

class ProfessionalAPI(AbstractView):
    
    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_professional(self, request):
        """Get a list of professionals

        Args:
            None
        """

        try:
            response_data = [{'code': v.name, 'name': v.value} for v in Professional.__members__.values()]

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class DestinationHistoryAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_destination_history(self, request):
        """Create a destination history

        Args:
            + user_code: String
            + country_code: String
            + city_id: int
            - district_id: int
            - ward_id: int
            - detail_address: String
            + start_time: String vd:'2000-01-26T01:23:45.123456Z'
            - end_time: String vd:'2000-01-26T01:23:45.123456Z'
            - note: String
        """

        accept_fields = [
            'user_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'start_time', 'end_time',
            'note',
        ]

        require_fields = [
            'user_code',
            'country_code', 'city_id',
            'start_time',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = DestinationHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'start_time', 'end_time',
            ])
            validator.extra_validate_to_create_destination_history()

            list_to_create_destination_history = [key for key in accepted_fields.keys()]
            list_to_create_destination_history = set(list_to_create_destination_history) - \
            {'user_code', 'country_code', 'city_id', 'district_id', 'ward_id',}
            list_to_create_destination_history = list(list_to_create_destination_history) + \
            [
                'user', 'country', 'city', 'district', 'ward',
            ]

            dict_to_create_destination_history = validator.get_data(list_to_create_destination_history)

            destination_history = DestinationHistory(**dict_to_create_destination_history)
            destination_history.save()

            serializer = DestinationHistorySerializer(destination_history, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_destination_history(self, request):
        """Get a destination history

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = DestinationHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_get_destination_history()

            destination_history = validator.get_field('destination_history')
            
            serializer = DestinationHistorySerializer(destination_history, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_destination_history(self, request):
        """Update a destination history

        Args:
            + id: int
            - country_code: String
            - city_id: int
            - district_id: int
            - ward_id: int
            - detail_address: String
            - start_time: String vd:'2000-01-26T01:23:45.123456Z'
            - end_time: String vd:'2000-01-26T01:23:45.123456Z'
            - note: String
        """

        accept_fields = [
            'id',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'start_time', 'end_time',
            'note',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = DestinationHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'start_time', 'end_time',
            ])
            validator.extra_validate_to_update_destination_history()

            list_to_update_destination_history = [key for key in accepted_fields.keys()]
            list_to_update_destination_history = set(list_to_update_destination_history) - \
            {'id', 'country_code', 'city_id', 'district_id', 'ward_id',}
            list_to_update_destination_history = list(list_to_update_destination_history) + \
            [
                'country', 'city', 'district', 'ward',
            ]

            dict_to_update_destination_history = validator.get_data(list_to_update_destination_history)

            destination_history = validator.get_field('destination_history')

            for attr, value in dict_to_update_destination_history.items(): 
                setattr(destination_history, attr, value)

            destination_history.save()

            serializer = DestinationHistorySerializer(destination_history, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='delete', detail=False)
    def delete_destination_history(self, request):
        """Delete a destination history

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = DestinationHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_delete_destination_history()

            destination_history = validator.get_field('destination_history')
            destination_history.delete()
            
            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_destination_history(self, request):
        """Get a list of destination history

        Args:
            + user_code: String
        """

        accept_fields = [
            'user_code',
        ]

        require_fields = [
            'user_code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = DestinationHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_filter_destination_history()

            list_to_filter_destination_history = [key for key in accepted_fields.keys()]

            dict_to_filter_destination_history = validator.get_data(list_to_filter_destination_history)

            dict_to_filter_destination_history.setdefault('order_by', 'start_time')

            filter = DestinationHistoryFilter(dict_to_filter_destination_history, queryset=DestinationHistory.objects.all())

            query_set = filter.qs
            query_set = query_set.select_related('user', 'country', 'city', 'district', 'ward')

            serializer = DestinationHistorySerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)
            
            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class QuarantineHistoryAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_quarantine_history(self, request):
        """Get a quarantine history

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = QuarantineHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_get_quarantine_history()

            quarantine_history = validator.get_field('quarantine_history')
            
            serializer = QuarantineHistorySerializer(quarantine_history, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_quarantine_history(self, request):
        """Get a list of quarantine history

        Args:
            + user_code: String
        """

        accept_fields = [
            'user_code',
        ]

        require_fields = [
            'user_code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = QuarantineHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_filter_quarantine_history()

            list_to_filter_quarantine_history = [key for key in accepted_fields.keys()]

            dict_to_filter_quarantine_history = validator.get_data(list_to_filter_quarantine_history)

            dict_to_filter_quarantine_history.setdefault('order_by', 'start_date')

            filter = QuarantineHistoryFilter(dict_to_filter_quarantine_history, queryset=QuarantineHistory.objects.all())

            query_set = filter.qs
            query_set = query_set.select_related('user', 'pandemic', 'quarantine_ward', 'quarantine_room__quarantine_floor__quarantine_building', 'created_by', 'updated_by')

            serializer = QuarantineHistorySerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)
            
            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='fix_invalid', detail=False)
    def fix_invalid_quarantine_history(self, request):
        """
        If this member is AVAILABLE:
            If this member does not have any quarantine history of pandemic of quarantine ward of this CustomUser, create a quarantine history
        If this member is LEAVE:
            If this member does not have any quarantine history, create a quarantine history

        Args:
            + user_code: String
        """

        accept_fields = [
            'user_code',
        ]

        require_fields = [
            'user_code',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = QuarantineHistoryValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_fix_empty_quarantine_history()

            user = validator.get_field('user')
            pandemic = user.quarantine_ward.pandemic

            if pandemic:
                if user.status == CustomUserStatus.AVAILABLE:
                    list_of_quarantine_history = list(QuarantineHistory.objects.filter(
                        user=user,
                        pandemic=pandemic,
                    ))
                    if len(list_of_quarantine_history) == 0:
                        quarantine_history = QuarantineHistory(
                            user=user,
                            pandemic=pandemic,
                            quarantine_ward=user.quarantine_ward,
                            quarantine_room=user.member_x_custom_user.quarantine_room,
                            status=QuarantineHistoryStatus.PRESENT,
                            start_date=user.member_x_custom_user.quarantined_at,
                            created_by=request.user,
                            updated_by=request.user,
                        )
                        quarantine_history.save()

                elif user.status == CustomUserStatus.LEAVE:
                    list_of_quarantine_history = list(QuarantineHistory.objects.filter(
                        user=user,
                    ))
                    if len(list_of_quarantine_history) == 0:
                        end_type = user.member_x_custom_user.quarantined_status if user.member_x_custom_user.quarantined_status in [MemberQuarantinedStatus.COMPLETED, MemberQuarantinedStatus.HOSPITALIZE] else None
                        quarantine_history = QuarantineHistory(
                            user=user,
                            pandemic=pandemic,
                            quarantine_ward=user.quarantine_ward,
                            quarantine_room=user.member_x_custom_user.quarantine_room,
                            status=QuarantineHistoryStatus.ENDED,
                            start_date=user.member_x_custom_user.quarantined_at,
                            end_date=user.member_x_custom_user.quarantined_finished_at,
                            end_type=end_type,
                            created_by=request.user,
                            updated_by=request.user,
                        )
                        quarantine_history.save()

            return self.response_handler.handle(message=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

def send_notification_is_last_test():
    try:
        pandemic = Pandemic.objects.get(name='Covid-19')
        day_between_tests = int(pandemic.day_between_tests)
    except:
        day_between_tests = int(os.environ.get('DAY_BETWEEN_TESTS', 5))

    last_tested_max = timezone.now() - datetime.timedelta(days=day_between_tests)

    dict_to_filter_need_test_members = {
        'role_name': 'MEMBER',
        'last_tested_max': last_tested_max,
        'status': CustomUserStatus.AVAILABLE,
    }

    filter = MemberFilter(dict_to_filter_need_test_members, queryset=CustomUser.objects.all())
    query_set = filter.qs
    
    create_and_send_noti_to_list_user(
        'Đến hạn xét nghiệm',
        'Theo quy định về lịch trình xét nghiệm, bạn đã đến hạn phải làm xét nghiệm',
        created_by=None,
        receive_user_list=list(query_set),
    )

class MemberAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    def get_background_disease_by_id(self, id):
        try:
            return BackgroundDisease.objects.get(id=id)
        except:
            return None

    def get_permissions(self):
        if self.action in ['register_member', 'bvdc_call_hospitalize_confirm']:
            self.permission_classes = [permissions.AllowAny]
        return super().get_permissions()

    def is_room_full(self, room):
        return room.member_x_quarantine_room.all().count() >= room.capacity

    def is_room_close(self, room, num_day_to_close_room):
        # All members in this room must have quarantined at >= 'num_day_to_close_room' day ago.
        time_now = timezone.now()
        for member in list(room.member_x_quarantine_room.all()):
            if member.quarantined_at and member.quarantined_at < time_now - datetime.timedelta(days=num_day_to_close_room):
                return True
        return False

    def count_members_same_label(self, room, label):
        if not label:
            return 0
        return room.member_x_quarantine_room.all().filter(label=label).count()

    def count_average_number_of_vaccine_doses(self, room):
        return_value = room.member_x_quarantine_room.all().aggregate(Avg('number_of_vaccine_doses'))['number_of_vaccine_doses__avg']
        if not return_value:
            # Nobody in room
            return_value = 1
        return return_value

    def count_members_same_gender(self, room, gender):
        if not gender:
            return 0
        return room.member_x_quarantine_room.all().filter(custom_user__gender=gender).count()

    def count_available_slot(self, room):
        return room.capacity - room.member_x_quarantine_room.all().count()

    def get_suitable_room_for_member(self, input_dict):
        """input_dict.keys() = [
            quarantine_ward
            gender
            label
            positive_test_now (can None)
            number_of_vaccine_doses
            old_quarantine_room (can None)
            not_quarantine_room_ids (can empty list)
        ]
        default, all field is not None

        """

        return_dict = dict()
        return_dict['room'] = None
        return_dict['warning'] = None

        # check old room
        if input_dict['old_quarantine_room'] and \
        input_dict['old_quarantine_room'].quarantine_floor.quarantine_building.quarantine_ward == input_dict['quarantine_ward'] and \
        input_dict['old_quarantine_room'].id not in input_dict['not_quarantine_room_ids']:
            if input_dict['positive_test_now'] in [False, None]:
                if self.count_positive_test_now_true_in_room(input_dict['old_quarantine_room']) == 0:
                    return_dict['room'] = input_dict['old_quarantine_room']
                    return_dict['warning'] = 'This room is current room of this member'
                    return return_dict
            else:
                if self.count_positive_test_now_not_true_in_room(input_dict['old_quarantine_room']) == 0:
                    return_dict['room'] = input_dict['old_quarantine_room']
                    return_dict['warning'] = 'This room is current room of this member'
                    return return_dict

        rooms = QuarantineRoom.objects.filter(quarantine_floor__quarantine_building__quarantine_ward = input_dict['quarantine_ward'])
        rooms = rooms.exclude(id__in=input_dict['not_quarantine_room_ids'])
        rooms = list(rooms)

        remain_rooms = [room for room in rooms if not self.is_room_full(room)]
        if len(remain_rooms) > 0:
            rooms = remain_rooms
        else:
            return_dict['room'] = None
            return_dict['warning'] = 'All rooms in this quarantine ward are full'
            return return_dict

        # check this room has positive / not positive member
        if input_dict['positive_test_now'] in [False, None]:
            remain_rooms = [room for room in rooms if self.count_positive_test_now_true_in_room(room) == 0]
        else:
            remain_rooms = [room for room in rooms if self.count_positive_test_now_not_true_in_room(room) == 0]
        if len(remain_rooms) > 0:
            rooms = remain_rooms
        else:
            return_dict['room'] = None
            return_dict['warning'] = 'All rooms in this quarantine ward are full or dont meet with this user positive_test_now'
            return return_dict

        tieu_chi = ['close_room', 'label', 'vaccine', 'gender', 'less_slot']
        # close_room: All members in this room must have quarantined at >= 'num_day_to_close_room' day ago.
        # label: This room must have at most number of members that is same label as this member.
        # vaccine: This room must have minimum average abs(this.number_of_vaccine_doses - another_member.number_of_vaccine_doses)
        # gender: This room must have at most number of members that is same gender as this member.
        # less_slot: This room must have at less number of available slot.

        # tieu_chi close_room
        if input_dict['positive_test_now'] in [False, None]:
            if input_dict['quarantine_ward'].pandemic:
                num_day_to_close_room = input_dict['quarantine_ward'].pandemic.num_day_to_close_room
            else:
                num_day_to_close_room = int(os.environ.get('NUM_DAY_TO_CLOSE_ROOM', 1))
            remain_rooms = [room for room in rooms if not self.is_room_close(room, num_day_to_close_room)]
            if len(remain_rooms) > 0:
                rooms = remain_rooms
            else:
                return_dict['room'] = None
                return_dict['warning'] = 'All rooms are not accept any more member'
                return return_dict

        # tieu_chi label
        count_each_room = [self.count_members_same_label(room, input_dict['label']) for room in rooms]
        max_same_label_in_room = max(count_each_room)
        remain_rooms = [rooms[i] for i in range(len(rooms)) if count_each_room[i] == max_same_label_in_room]
        rooms = remain_rooms
        
        # tieu_chi vaccine
        difference_each_room = [abs(self.count_average_number_of_vaccine_doses(room) - input_dict['number_of_vaccine_doses']) for room in rooms]
        min_difference_each_room = min(difference_each_room)
        remain_rooms = [rooms[i] for i in range(len(rooms)) if difference_each_room[i] == min_difference_each_room]
        rooms = remain_rooms

        # tieu_chi gender
        count_each_room = [self.count_members_same_gender(room, input_dict['gender']) for room in rooms]
        max_same_gender_in_room = max(count_each_room)
        remain_rooms = [rooms[i] for i in range(len(rooms)) if count_each_room[i] == max_same_gender_in_room]
        rooms = remain_rooms

        # tieu_chi less_slot
        count_each_room = [self.count_available_slot(room) for room in rooms]
        min_available_slot = min(count_each_room)
        remain_rooms = [rooms[i] for i in range(len(rooms)) if count_each_room[i] == min_available_slot]
        rooms = remain_rooms

        return_dict['room'] = rooms[0]
        return_dict['warning'] = None
        return return_dict

    def check_room_for_member(self, user, room):
        # Check if this member can be set to this room
        if not hasattr(user, 'member_x_custom_user') or user.role.name != 'MEMBER':
            return messages.ISNOTMEMBER
        if room.quarantine_floor.quarantine_building.quarantine_ward != user.quarantine_ward:
            return 'This room is not in the quarantine ward of this user'
        if user.member_x_custom_user.quarantine_room == room:
            return messages.SUCCESS
        if self.is_room_full(room):
            return 'This room is full'
        if user.member_x_custom_user.positive_test_now == True:
            if self.count_positive_test_now_not_true_in_room(room) >= 1:
                return 'This member positive, but this room has member that is not positive'
        else:
            if self.count_positive_test_now_true_in_room(room) >= 1:
                return 'This room has member that is positive'

            # check is room close (not accept any more member)
            quarantine_ward = room.quarantine_floor.quarantine_building.quarantine_ward
            if quarantine_ward.pandemic:
                num_day_to_close_room = quarantine_ward.pandemic.num_day_to_close_room
            else:
                num_day_to_close_room = int(os.environ.get('NUM_DAY_TO_CLOSE_ROOM', 1))
            if self.is_room_close(room, num_day_to_close_room):
                return 'This room is close (not accept any more member)'
        return messages.SUCCESS

    def get_suitable_care_staff_for_member(self, input_dict):
        """input_dict.keys() = [
            quarantine_floor
        ]
        default, all field is not None
        """

        return_dict = dict()
        return_dict['care_staff'] = None
        return_dict['warning'] = None

        query_set = Staff.objects.filter(
            care_area__icontains=input_dict['quarantine_floor'].id,
            custom_user__status=CustomUserStatus.AVAILABLE,
            custom_user__quarantine_ward=input_dict['quarantine_floor'].quarantine_building.quarantine_ward,
        )
        if query_set.count() == 0:
            query_set = Staff.objects.filter(
                custom_user__status=CustomUserStatus.AVAILABLE,
                custom_user__quarantine_ward=input_dict['quarantine_floor'].quarantine_building.quarantine_ward,
            )
        try:
            return_dict['care_staff'] = query_set.annotate(num_care_member=Count('custom_user__member_x_care_staff')).order_by('num_care_member')[:1].get().custom_user
        except Exception as exception:
            return_dict['care_staff'] = None
            return_dict['warning'] = 'Cannot set care_staff for this member'
        return return_dict

    def count_positive_test_now_true_in_room(self, room):
        if room:
            return room.member_x_quarantine_room.all().filter(positive_test_now=True).count()
        return 0

    def count_positive_test_now_not_true_in_room(self, room):
        if room:
            return room.member_x_quarantine_room.all().filter(Q(positive_test_now=False) | Q(positive_test_now__isnull=True)).count()
        return 0

    def do_after_change_room_of_member_work(self, member, old_room):
        """
        run this function after moving a member from null room to a room, old room to new room or a room to out quarantine;
        old_room can be None, mean not have old_room;
        new_room is member.quarantine_room;
        """
        new_room = member.quarantine_room

        # old room
        if old_room != None and new_room != old_room:
            if member.positive_test_now == True:
                remain_members_in_old_room = old_room.member_x_quarantine_room.all().exclude(id=member.id)
                number_of_remain_positive_member_in_old_room = remain_members_in_old_room.filter(positive_test_now=True).count()
                if number_of_remain_positive_member_in_old_room == 0:
                    quarantine_ward = old_room.quarantine_floor.quarantine_building.quarantine_ward
                    for each_member in remain_members_in_old_room:
                        if each_member.label != MemberLabel.F0:
                            if quarantine_ward.pandemic:
                                if each_member.number_of_vaccine_doses < 2:
                                    remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_not_vac
                                else:
                                    remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_vac
                            else:
                                if each_member.number_of_vaccine_doses < 2:
                                    remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_NOT_VAC', 14))
                                else:
                                    remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_VAC', 10))
                            old_quarantined_finish_expected_at = each_member.quarantined_finish_expected_at
                            new_quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                            if not old_quarantined_finish_expected_at or old_quarantined_finish_expected_at < new_quarantined_finish_expected_at:
                                each_member.quarantined_finish_expected_at = new_quarantined_finish_expected_at
                                each_member.save()
        
        # new room
        if new_room != None and new_room != old_room:
            if member.label != MemberLabel.F0:
                all_members_in_new_room = list(new_room.member_x_quarantine_room.all().exclude(id=member.id))
                all_members_in_new_room.append(member)

                # set quarantined_finish_expected_at for all member in new room (include this member)
                number_of_other_members_in_new_room = new_room.member_x_quarantine_room.all().exclude(id=member.id).count()
                if number_of_other_members_in_new_room >= 1:
                    quarantine_ward = new_room.quarantine_floor.quarantine_building.quarantine_ward
                    for each_member in all_members_in_new_room:
                        if quarantine_ward.pandemic:
                            if each_member.number_of_vaccine_doses < 2:
                                remain_qt = quarantine_ward.pandemic.remain_qt_cc_not_pos_not_vac
                            else:
                                remain_qt = quarantine_ward.pandemic.remain_qt_cc_not_pos_vac
                        else:
                            if each_member.number_of_vaccine_doses < 2:
                                remain_qt = int(os.environ.get('REMAIN_QT_CC_NOT_POS_NOT_VAC', 7))
                            else:
                                remain_qt = int(os.environ.get('REMAIN_QT_CC_NOT_POS_VAC', 5))
                        old_quarantined_finish_expected_at = each_member.quarantined_finish_expected_at
                        new_quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                        if not old_quarantined_finish_expected_at or old_quarantined_finish_expected_at < new_quarantined_finish_expected_at:
                            each_member.quarantined_finish_expected_at = new_quarantined_finish_expected_at
                            each_member.save()

                # set label
                label_tool = LabelTool()
                most_serious_label = member.label
                for each_member in all_members_in_new_room:
                    if label_tool.compare_label(each_member.label, most_serious_label) == 1:
                        most_serious_label = each_member.label
                if 0 <= label_tool.get_value_of_label(most_serious_label) <= 2:
                    down_label = label_tool.down_label(most_serious_label)
                    for each_member in all_members_in_new_room:
                        if label_tool.compare_label(each_member.label, down_label) == -1:
                            each_member.label = down_label
                            each_member.save()

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='register', detail=False)
    def register_member(self, request):
        """For someone outside to register quarantine

        Args:
            + phone_number: String
            + password: String
            + quarantine_ward_id: int
        """

        accept_fields = [
            'phone_number', 'password',
            'quarantine_ward_id', 
        ]

        require_fields = [
            'phone_number', 'password',
            'quarantine_ward_id', 
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'phone_number', 'password',
            ])

            validator.extra_validate_to_register_member()

            # create CustomUser

            list_to_create_custom_user = ['phone_number', 'quarantine_ward']

            dict_to_create_custom_user = validator.get_data(list_to_create_custom_user)

            custom_user = CustomUser(**dict_to_create_custom_user)
            password = accepted_fields['password']
            custom_user.set_password(password)
            custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            while (validator.is_code_exist(custom_user.code)):
                custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            
            custom_user.role = Role.objects.get(name='MEMBER')
            custom_user.status = CustomUserStatus.WAITING

            # create Member

            member = Member()
            member.custom_user = custom_user

            custom_user.save()
            member.save()

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            member_serializer = MemberSerializer(member, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['member'] = member_serializer.data
            
            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_member(self, request):
        """Create a member

        Args:
            + full_name: String
            + phone_number: String
            - email: String
            + birthday: String 'dd/mm/yyyy'
            + gender: String ['MALE', 'FEMALE']
            + nationality_code: String
            + country_code: String
            + city_id: int
            + district_id: int
            + ward_id: int
            + detail_address: String
            - professional: String in Professional enum, call api to get list of professionals
            - health_insurance_number: String
            + identity_number: String
            - passport_number: String
            + quarantine_ward_id: int
            - quarantine_room_id: int
            - label: String ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            - first_positive_test_date: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_reason: String
            - quarantined_at: String vd:'2000-01-26T01:23:45.123456Z'
            - positive_tested_before: boolean
            - background_disease: String '<id>,<id>,<id>'
            - other_background_disease: String
            - number_of_vaccine_doses: int
            - care_staff_code: String
        """

        accept_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'professional',
            'quarantine_ward_id', 'quarantine_room_id',
            'label', 'first_positive_test_date',
            'quarantine_reason',
            'quarantined_at', 'positive_tested_before',
            'background_disease', 'other_background_disease',
            'number_of_vaccine_doses', 'care_staff_code',
        ]

        require_fields = [
            'full_name', 'phone_number',
            'birthday', 'gender',
            'identity_number', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'quarantine_ward_id',
        ]

        custom_user_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'professional', 'quarantine_ward_id',
        ]

        member_fields = [
            'quarantine_room_id', 'quarantine_reason',
            'label', 'first_positive_test_date',
            'quarantined_at', 'positive_tested_before',
            'background_disease', 'other_background_disease',
            'number_of_vaccine_doses', 'care_staff_code',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'phone_number', 'email', 'birthday', 'gender',
                'passport_number', 'health_insurance_number', 'identity_number',
                'professional',
                'label', 'first_positive_test_date',
                'quarantined_at', 'positive_tested_before',
                'background_disease', 'number_of_vaccine_doses',
            ])
            validator.extra_validate_to_create_member()

            # create CustomUser

            list_to_create_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_create_custom_user = set(list_to_create_custom_user) - \
            {'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_create_custom_user = list(list_to_create_custom_user) + \
            [
                'nationality', 'country', 'city', 'district', 'ward',
                'quarantine_ward',
            ]

            dict_to_create_custom_user = validator.get_data(list_to_create_custom_user)

            custom_user = CustomUser(**dict_to_create_custom_user)
            custom_user.set_password('123456')
            custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            while (validator.is_code_exist(custom_user.code)):
                custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            custom_user.created_by = request.user
            custom_user.updated_by = request.user
            custom_user.role = Role.objects.get(name='MEMBER')

            # create Member

            list_to_create_member = [key for key in accepted_fields.keys() if key in member_fields]
            list_to_create_member = set(list_to_create_member) - \
            {'quarantine_room_id', 'care_staff_code'}
            list_to_create_member = list(list_to_create_member) + \
            ['quarantined_at', 'quarantined_finish_expected_at', 'positive_test_now', 'care_staff',]

            dict_to_create_member = validator.get_data(list_to_create_member)

            member = Member(**dict_to_create_member)
            member.custom_user = custom_user

            # extra set room for this member
            if hasattr(validator, '_quarantine_room'):
                # this field is received and not None
                quarantine_room = validator.get_field('quarantine_room')
                check_room_result = self.check_room_for_member(custom_user, quarantine_room)
                if check_room_result != messages.SUCCESS:
                    raise exceptions.ValidationException({'quarantine_room_id': check_room_result})
            else:
                input_dict_for_get_suitable_room = dict()
                input_dict_for_get_suitable_room['quarantine_ward'] = custom_user.quarantine_ward
                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                input_dict_for_get_suitable_room['label'] = member.label
                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                input_dict_for_get_suitable_room['old_quarantine_room'] = None
                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                quarantine_room = suitable_room_dict['room']
                warning = suitable_room_dict['warning']
                if not quarantine_room:
                    raise exceptions.ValidationException({'main': warning})

            member.quarantine_room = quarantine_room
            member.number_of_vaccine_doses = 0

            # create QuarantineHistory
            quarantine_history = QuarantineHistory(
                user=custom_user,
                pandemic=custom_user.quarantine_ward.pandemic,
                quarantine_ward=custom_user.quarantine_ward,
                quarantine_room=member.quarantine_room,
                status=QuarantineHistoryStatus.PRESENT,
                start_date=member.quarantined_at,
                created_by=request.user,
                updated_by=request.user,
            )

            custom_user.save()
            member.save()

            quarantine_history.save()

            self.do_after_change_room_of_member_work(member, None)

            member = Member.objects.get(id=member.id)

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            member_serializer = MemberSerializer(member, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['member'] = member_serializer.data
            
            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)
    
    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create_by_file', detail=False)
    def create_member_by_file(self, request):
        """Create a member by file (csv, xlsx)

        Args:
            + file: .csv, .xlsx
        """

        gender_switcher = {
            "Nam": 'MALE',
            "Nữ": 'FEMALE',
        }

        boolean_switcher = {
            "Có": True,
            "Không": False
        }

        custom_user_fields = [
            'full_name', 'email', 'phone_number',
            'birthday', 'nationality_code', 'gender',
            'country_code', 'city_id', 'district_id', 'ward_id', 'detail_address', 
            'identity_number', 'health_insurance_number', 'passport_number',
            'quarantine_ward_id',
        ]

        member_fields = [
            'quarantine_room_id',
            'label', 'quarantined_at', 'positive_tested_before',
            'background_disease', 'other_background_disease',
            'number_of_vaccine_doses', 'care_staff_code',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            file_name = request.FILES.get('file', False)
            if file_name:
                extension = os.path.splitext(str(file_name))[1]
                iter_rows = None
                error = dict()
                custom_user_ids = list()
                is_csv_file = False

                if (extension == ".xlsx"):
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb["Sheet1"]
                    iter_rows = ws.iter_rows(min_row=2, min_col=2, max_col=25)
                elif (extension == ".csv"):
                    iter_rows = csv.DictReader(codecs.iterdecode(file_name, encoding='utf-8'))
                    is_csv_file = True

                for row_index, row in enumerate(iter_rows):
                    try:
                        dict_custom_user_data = dict()
                        dict_member_data = dict()
                        is_available = False
                        check_data = None

                        if is_csv_file: 
                            row = list(row.values())[1:]
                            check_data = row[0]
                        else:
                            check_data = row[0].value
                        if check_data in [None, '']:
                            continue

                        for index, data in enumerate(row):
                            value = data.value if not is_csv_file else data
                            if (index < 15):
                                if index == 3:
                                    if (value != None):
                                        if not is_csv_file: 
                                            value = value.strftime('%d/%m/%Y')
                                        else:
                                            value = str(value)
                                            value = str(datetime.datetime.strptime(value, '%d/%m/%Y').strftime('%d/%m/%Y'))
                                elif index == 5:
                                    value = str(value)
                                    value = gender_switcher[value]
                                elif index in [7, 8, 9, 14]:
                                    value = int(value)
                                else:
                                    value = str(value)
                                dict_custom_user_data[custom_user_fields[index]] = value

                            elif (index == 15):

                                if str(value) == "Có":
                                    is_available = True

                            else:
                                if (index == 18):
                                    if (value != None):
                                        if not is_csv_file:
                                            value = value.strftime('%Y-%m-%dT%H:%M:%S.%f%zZ')
                                        else:
                                            value = str(value)
                                            value = str(datetime.datetime.strptime(value, '%d/%m/%Y').strftime('%Y-%m-%dT%H:%M:%S.%f%zZ'))
                                    
                                if value in ["Có", "Không"]:
                                    value = str(value)
                                    value = boolean_switcher[value]
                                if value != None or value != "":
                                    dict_member_data[member_fields[index - 16]] = value
                            
                        user_validator = UserValidator(**dict_custom_user_data)
                        user_validator.is_valid_fields([
                            'phone_number', 'email', 'birthday', 'gender',
                            'passport_number', 'health_insurance_number', 'identity_number',
                        ])
                        user_validator.extra_validate_to_create_member()

                        # create CustomUser
                        custom_user_field = set(custom_user_fields) - \
                        {'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
                        custom_user_field = list(custom_user_field) + \
                        [
                            'nationality', 'country', 'city', 'district', 'ward',
                            'quarantine_ward',
                        ]

                        dict_to_create_custom_user = user_validator.get_data(custom_user_field)

                        custom_user = CustomUser(**dict_to_create_custom_user)
                        custom_user.set_password('123456')
                        custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
                        while (user_validator.is_code_exist(custom_user.code)):
                            custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
                        custom_user.created_by = request.user
                        custom_user.updated_by = request.user
                        custom_user.role = Role.objects.get(name='MEMBER')
                        if not is_available:
                            dict_member_data.pop("quarantine_room")
                            custom_user.status = CustomUserStatus.WAITING
                        
                        # create Member
                        dict_member_data["quarantine_ward"] = custom_user.quarantine_ward
                        
                        member_validator = UserValidator(**dict_member_data)
                        member_validator.is_valid_fields([
                            'label', 'quarantined_at', 'positive_tested_before',
                            'background_disease', 'number_of_vaccine_doses',
                        ])
                        member_validator.extra_validate_to_create_member()
                
                        member_field = set(member_fields) - \
                        {'quarantine_room_id', 'care_staff_code'}
                        member_field = list(member_field) + \
                        ['quarantined_at', 'quarantined_finish_expected_at', 'positive_test_now', 'care_staff',]

                        dict_to_create_member = member_validator.get_data(member_field)

                        member = Member(**dict_to_create_member)
                        member.custom_user = custom_user

                        # extra set room for this member
                        if is_available:
                            if hasattr(member_validator, '_quarantine_room'):
                                # this field is received and not None
                                quarantine_room = member_validator.get_field('quarantine_room')
                                check_room_result = self.check_room_for_member(custom_user, quarantine_room)
                                if check_room_result != messages.SUCCESS:
                                    raise exceptions.ValidationException({'quarantine_room_id': check_room_result})
                            else:
                                input_dict_for_get_suitable_room = dict()
                                input_dict_for_get_suitable_room['quarantine_ward'] = custom_user.quarantine_ward
                                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                                input_dict_for_get_suitable_room['label'] = member.label
                                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                                input_dict_for_get_suitable_room['old_quarantine_room'] = None
                                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                                quarantine_room = suitable_room_dict['room']
                                warning = suitable_room_dict['warning']
                                if not quarantine_room:
                                    raise exceptions.ValidationException({'main': warning})

                            member.quarantine_room = quarantine_room
                            member.number_of_vaccine_doses = 0

                            # create QuarantineHistory
                            quarantine_history = QuarantineHistory(
                                user=custom_user,
                                pandemic=custom_user.quarantine_ward.pandemic,
                                quarantine_ward=custom_user.quarantine_ward,
                                quarantine_room=member.quarantine_room,
                                status=QuarantineHistoryStatus.PRESENT,
                                start_date=member.quarantined_at,
                                created_by=request.user,
                                updated_by=request.user,
                            )

                        custom_user.save()
                        member.save()

                        if is_available:
                            quarantine_history.save()
                            self.do_after_change_room_of_member_work(member, None)

                        custom_user_ids += [custom_user.id]
                    except Exception as e:
                        error[str(row_index + 1)] = str(e)
                        pass
                
                custom_user_data = CustomUser.objects.filter(id__in=custom_user_ids)
                member_serializer = FilterMemberSerializer(custom_user_data, many=True)
                response_data = dict()
                response_data["user_success"] = member_serializer.data
                response_data["user_fail"] = error
                
            else:
                raise exceptions.InvalidArgumentException({'main': messages.FILE_IMPORT_EMPTY})

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_user(self, request):
        """Get a user, if dont get id, will return user sending request

        Args:
            - code: int
        """

        accept_fields = [
            'code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)

            custom_user = None

            if 'code' in accepted_fields.keys():
                if validator.is_code_exist():
                    custom_user = validator.get_field('custom_user')
                else:
                    raise exceptions.NotFoundException({'code': messages.NOT_EXIST})
            else:
                custom_user = request.user
            
            response_data = dict()

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)

            response_data['custom_user'] = custom_user_serializer.data

            if hasattr(custom_user, 'member_x_custom_user'):
                member = custom_user.member_x_custom_user
                member_serializer = MemberSerializer(member, many=False)

                response_data['member'] = member_serializer.data

            if hasattr(custom_user, 'manager_x_custom_user'):
                manager = custom_user.manager_x_custom_user
                manager_serializer = ManagerSerializer(manager, many=False)

                response_data['manager'] = manager_serializer.data

            if hasattr(custom_user, 'staff_x_custom_user'):
                staff = custom_user.staff_x_custom_user
                staff_serializer = StaffSerializer(staff, many=False)

                response_data['staff'] = staff_serializer.data
            
            return self.response_handler.handle(data=response_data)
        
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get_by_phone_number', detail=False)
    def get_user_by_phone_number(self, request):
        """Get a user by phone number

        Args:
            + phone_number: String
        """

        accept_fields = [
            'phone_number',
        ]

        require_fields = [
            'phone_number',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_get_user_by_phone_number()

            custom_user = validator.get_field('custom_user')
            
            response_data = dict()
            response_data['code'] = custom_user.code
            response_data['full_name'] = custom_user.full_name
            
            return self.response_handler.handle(data=response_data)
        
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get_timeline', detail=False)
    def get_member_timeline(self, request):
        """Get a member timeline

        Args:
            + code: String
        """

        accept_fields = [
            'code',
        ]

        require_fields = [
            'code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_get_member_timeline()

            custom_user = validator.get_field('custom_user')

            if request.user.role.name == 'MEMBER' and request.user != custom_user:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            response_data = dict()

            if custom_user.member_x_custom_user.quarantined_at:
                vntz = pytz.timezone('Asia/Saigon')
                start_time = custom_user.member_x_custom_user.quarantined_at.astimezone(vntz) - datetime.timedelta(14)
                if custom_user.member_x_custom_user.quarantined_finished_at:
                    end_time = custom_user.member_x_custom_user.quarantined_finished_at.astimezone(vntz)
                else:
                    end_time = timezone.now().astimezone(vntz)
                end_time = datetime.datetime(end_time.year, end_time.month, end_time.day, 23, 59, 59, 999999).astimezone(vntz)

                # vaccine dose
                vaccine_dose_query_set = VaccineDose.objects.filter(
                    custom_user=custom_user,
                    injection_date__isnull=False,
                )
                vaccine_dose_query_set = vaccine_dose_query_set.order_by('injection_date')
                vaccine_dose_query_set = vaccine_dose_query_set.select_related('custom_user', 'vaccine')
                vaccine_dose_list = BaseVaccineDoseSerializer(vaccine_dose_query_set, many=True).data

                new_vaccine_dose_list = []
                for item in vaccine_dose_list:
                    new_item = dict()
                    new_item['type'] = 'vaccine_dose'
                    new_item['data'] = item
                    new_vaccine_dose_list += [new_item]
                vaccine_dose_list = new_vaccine_dose_list

                # destination history
                destination_history_query_set = DestinationHistory.objects.filter(
                    user=custom_user,
                    start_time__gte=start_time,
                    start_time__lte=end_time,
                )
                destination_history_query_set = destination_history_query_set.order_by('start_time')
                destination_history_query_set = destination_history_query_set.select_related('user', 'ward__district__city__country', 'district__city__country', 'city__country', 'country')
                destination_history_list = DestinationHistorySerializer(destination_history_query_set, many=True).data

                new_destination_history_list = []
                for item in destination_history_list:
                    new_item = dict()
                    new_item['type'] = 'destination_history'
                    new_item['data'] = item
                    new_destination_history_list += [new_item]
                destination_history_list = new_destination_history_list

                # quarantine history
                quarantine_history_query_set = QuarantineHistory.objects.filter(
                    user=custom_user,
                    start_date__gte=start_time,
                    start_date__lte=end_time,
                )
                quarantine_history_query_set = quarantine_history_query_set.order_by('start_date')
                quarantine_history_query_set = quarantine_history_query_set.select_related('user', 'quarantine_ward', 'quarantine_room__quarantine_floor__quarantine_building')
                quarantine_history_list = BaseQuarantineHistorySerializer(quarantine_history_query_set, many=True).data
                if not quarantine_history_list:
                    raise exceptions.ValidationException({'quarantine_history': messages.INVALID})
                last_quarantine_history = quarantine_history_list.pop()
                new_quarantine_history_list = [last_quarantine_history]
                while quarantine_history_list:
                    last_quarantine_history = quarantine_history_list.pop()
                    if last_quarantine_history['end_type'] == QuarantineHistoryEndType.CHANGE_ROOM:
                        new_quarantine_history_list.insert(0, last_quarantine_history)
                    else:
                        break
                
                quarantine_history_list = []
                new_quarantine_history_list[0]['type'] = 'start_quarantine'
                if len(new_quarantine_history_list) >= 2:
                    for item in new_quarantine_history_list[1:]:
                        item['type'] = 'change_room'
                last_quarantine_history = dict()
                is_have_last = True
                if custom_user.status == CustomUserStatus.LEAVE:
                    last_quarantine_history['type'] = custom_user.member_x_custom_user.quarantined_status
                    last_quarantine_history['start_date'] = str(custom_user.member_x_custom_user.quarantined_finished_at.astimezone(vntz))
                    last_quarantine_history['start_date'] = last_quarantine_history['start_date'][:10] + 'T' + last_quarantine_history['start_date'][11:]
                    last_quarantine_history['note'] = new_quarantine_history_list[-1]['note']
                    last_quarantine_history['quarantine_ward'] = None
                    last_quarantine_history['quarantine_building'] = None
                    last_quarantine_history['quarantine_floor'] = None
                    last_quarantine_history['quarantine_room'] = None
                elif custom_user.member_x_custom_user.quarantined_finish_expected_at:
                    last_quarantine_history['type'] = 'expect_finish'
                    last_quarantine_history['start_date'] = str(custom_user.member_x_custom_user.quarantined_finish_expected_at.astimezone(vntz))
                    last_quarantine_history['start_date'] = last_quarantine_history['start_date'][:10] + 'T' + last_quarantine_history['start_date'][11:]
                    last_quarantine_history['note'] = None
                    last_quarantine_history['quarantine_ward'] = None
                    last_quarantine_history['quarantine_building'] = None
                    last_quarantine_history['quarantine_floor'] = None
                    last_quarantine_history['quarantine_room'] = None
                else:
                    is_have_last = False
                for item in new_quarantine_history_list:
                    item['note'] = None
                if is_have_last:
                    new_quarantine_history_list += [last_quarantine_history]
                
                for item in new_quarantine_history_list:
                    new_item = dict()
                    new_item['type'] = 'quarantine_history'
                    new_item['data'] = {
                        'type': item['type'],
                        'start_date': item['start_date'],
                        'quarantine_ward': item['quarantine_ward'],
                        'quarantine_building': item['quarantine_building'],
                        'quarantine_floor': item['quarantine_floor'],
                        'quarantine_room': item['quarantine_room'],
                        'note': item['note'],
                    }
                    quarantine_history_list += [new_item]

                # medical declaration
                medical_declaration_query_set = MedicalDeclaration.objects.filter(
                    user=custom_user,
                    created_at__gte=start_time,
                    created_at__lte=end_time,
                )
                medical_declaration_query_set = medical_declaration_query_set.order_by('created_at')
                medical_declaration_query_set = medical_declaration_query_set.select_related('user')
                medical_declaration_list = BaseMedicalDeclarationSerializer(medical_declaration_query_set, many=True).data
                
                new_medical_declaration_list = []
                for item in medical_declaration_list:
                    new_item = dict()
                    new_item['type'] = 'medical_declaration'
                    new_item['data'] = item
                    new_medical_declaration_list += [new_item]
                medical_declaration_list = new_medical_declaration_list

                # test
                test_query_set = Test.objects.filter(
                    user=custom_user,
                    result__in=[TestResult.POSITIVE, TestResult.NEGATIVE],
                    created_at__gte=start_time,
                    created_at__lte=end_time,
                )
                test_query_set = test_query_set.order_by('created_at')
                test_query_set = test_query_set.select_related('user__member_x_custom_user')
                test_list = BaseTestSerializer(test_query_set, many=True).data

                new_test_list = []
                for item in test_list:
                    new_item = dict()
                    new_item['type'] = 'test'
                    new_item['data'] = item
                    new_test_list += [new_item]
                test_list = new_test_list

                list_of_all = vaccine_dose_list + destination_history_list + quarantine_history_list + medical_declaration_list + test_list
                def key_created_at(a):
                    if a['type'] == 'destination_history':
                        return a['data']['start_time']
                    elif a['type'] == 'quarantine_history':
                        return a['data']['start_date']
                    elif a['type'] == 'vaccine_dose':
                        return a['data']['injection_date']
                    else:
                        return a['data']['created_at']
                list_of_all.sort(key=key_created_at)

                for item in list_of_all:
                    if item['type'] == 'destination_history':
                        date = str(item['data']['start_time'])[:10]
                    elif item['type'] == 'quarantine_history':
                        date = str(item['data']['start_date'])[:10]
                    elif item['type'] == 'vaccine_dose':
                        date = str(item['data']['injection_date'])[:10]
                    else:
                        date = str(item['data']['created_at'])[:10]
                    
                    if date in response_data.keys():
                        response_data[date] += [item]
                    else:
                        response_data[date] = [item]

            else:
                raise exceptions.ValidationException({'quarantined_at': messages.EMPTY})
            
            return self.response_handler.handle(data=response_data)
        
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_member(self, request):
        """Update a member, if dont get code, will update member sending request

        Args:
            - code: int
            - full_name: String
            - email: String
            - birthday: String 'dd/mm/yyyy'
            - gender: String ['MALE', 'FEMALE']
            - professional: String in Professional enum, call api to get list of professionals
            - nationality_code: String
            - country_code: int
            - city_id: int
            - district_id: int
            - ward_id: int
            - detail_address: String
            - health_insurance_number: String
            - identity_number: String
            - passport_number: String
            - quarantine_ward_id: int
            - quarantine_room_id: int
            - label: String ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            - first_positive_test_date: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_reason: String
            - quarantined_at: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_finish_expected_at: String vd:'2000-01-26T01:23:45.123456Z'
            - positive_tested_before: boolean
            - background_disease: String '<id>,<id>,<id>'
            - other_background_disease: String
            - care_staff_code: String
        """

        accept_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id', 'quarantine_room_id',
            'label', 'first_positive_test_date', 'professional',
            'quarantine_reason',
            'quarantined_at', 'quarantined_finish_expected_at',
            'positive_tested_before',
            'background_disease', 'other_background_disease',
            'care_staff_code',
        ]

        custom_user_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender',
            'professional', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        member_fields = [
            'quarantine_room_id', 'label',
            'first_positive_test_date', 'quarantine_reason',
            'quarantined_at', 'quarantined_finish_expected_at',
            'positive_tested_before',
            'background_disease', 'other_background_disease',
            'care_staff_code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            if 'code' not in accepted_fields.keys():
                accepted_fields['code'] = request.user.code

            validator = UserValidator(**accepted_fields)
            validator.is_valid_fields([
                'email', 'birthday', 'gender', 'passport_number',
                'health_insurance_number', 'identity_number',
                'label', 'first_positive_test_date', 'professional',
                'quarantined_at', 'quarantined_finish_expected_at',
                'positive_tested_before',
                'background_disease',
            ])
            validator.extra_validate_to_update_member(sender=request.user)

            # update CustomUser

            custom_user = validator.get_field('custom_user')

            if request.user.role.name == 'MEMBER' and request.user != custom_user:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            list_to_update_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_update_custom_user = set(list_to_update_custom_user) - \
            {'code', 'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_update_custom_user = list(list_to_update_custom_user) + \
            ['nationality', 'country', 'city', 'district', 'ward', 'quarantine_ward']
            dict_to_update_custom_user = validator.get_data(list_to_update_custom_user)

            for attr, value in dict_to_update_custom_user.items(): 
                setattr(custom_user, attr, value)

            custom_user.updated_by = request.user

            # update Member
            member = custom_user.member_x_custom_user

            list_to_update_member = [key for key in accepted_fields.keys() if key in member_fields]
            list_to_update_member = set(list_to_update_member) - \
            {'quarantine_room_id', 'label', 'care_staff_code',}
            list_to_update_member = list(list_to_update_member) + \
            ['quarantined_finish_expected_at', 'care_staff',]

            dict_to_update_member = validator.get_data(list_to_update_member)

            for attr, value in dict_to_update_member.items(): 
                setattr(member, attr, value)

            # because extra_validate func, label and room cannot change together

            # check label
            old_label = member.label
            new_label = validator.get_field('label')
            if new_label:
                if new_label != old_label:
                    if request.user.role.name == 'MEMBER' and custom_user.status == CustomUserStatus.AVAILABLE:
                        raise exceptions.AuthenticationException({'label': messages.NO_PERMISSION})

                    member.label = new_label

                    if old_label != MemberLabel.F0 and new_label == MemberLabel.F0:
                        member.positive_test_now = True
                        
                        if member.custom_user.status == CustomUserStatus.AVAILABLE:
                            if member.quarantined_finish_expected_at == None:
                                quarantine_ward = custom_user.quarantine_ward
                                if quarantine_ward.pandemic:
                                    if member.number_of_vaccine_doses < 2:
                                        remain_qt = quarantine_ward.pandemic.remain_qt_pos_not_vac
                                    else:
                                        remain_qt = quarantine_ward.pandemic.remain_qt_pos_vac
                                else:
                                    if member.number_of_vaccine_doses < 2:
                                        remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                                    else:
                                        remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                                member.quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)

                            # affect other member in this room
                            this_room = member.quarantine_room
                            other_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=member.id)
                            for each_member in list(other_members_in_this_room):
                                if each_member.label != MemberLabel.F0:
                                    each_member.label = MemberLabel.F1
                                    each_member.quarantined_finish_expected_at = None
                                    each_member.save()

                    elif old_label == MemberLabel.F0 and new_label != MemberLabel.F0:
                        member.positive_test_now = None

                        if member.custom_user.status == CustomUserStatus.AVAILABLE:
                            # check if this room have any positive member
                            this_room = member.quarantine_room
                            number_of_other_positive_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=member.id).filter(positive_test_now=True).count()
                            if number_of_other_positive_members_in_this_room >= 1:
                                member.label = MemberLabel.F1
                                member.quarantined_finish_expected_at = None
                            else:
                                # affect other member in this room
                                quarantine_ward = this_room.quarantine_floor.quarantine_building.quarantine_ward
                                other_members_in_this_room = list(this_room.member_x_quarantine_room.all().exclude(id=member.id))
                                for each_member in other_members_in_this_room:
                                    if each_member.label != MemberLabel.F0:
                                        if not each_member.quarantined_finish_expected_at:
                                            if quarantine_ward.pandemic:
                                                if each_member.number_of_vaccine_doses < 2:
                                                    remain_qt = quarantine_ward.pandemic.quarantine_time_not_vac
                                                else:
                                                    remain_qt = quarantine_ward.pandemic.quarantine_time_vac
                                            else:
                                                if each_member.number_of_vaccine_doses < 2:
                                                    remain_qt = int(os.environ.get('QUARANTINE_TIME_NOT_VAC', 14))
                                                else:
                                                    remain_qt = int(os.environ.get('QUARANTINE_TIME_VAC', 10))
                                            each_member.quarantined_finish_expected_at = each_member.quarantined_at + datetime.timedelta(days=remain_qt)
                                            each_member.save()

            # check room
            old_room = member.quarantine_room
            new_room = validator.get_field('quarantine_room')
            if new_room:
                if new_room != old_room:
                    if request.user.role.name == 'MEMBER':
                        raise exceptions.AuthenticationException({'quarantine_room_id': messages.NO_PERMISSION})

                    check_room_result = self.check_room_for_member(user=custom_user, room=new_room)
                    if check_room_result != messages.SUCCESS:
                        raise exceptions.ValidationException({'quarantine_room_id': check_room_result})
                    member.quarantine_room = new_room

                    # update QuarantineHistory
                    old_present_quarantine_history = list(QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT))
                    if len(old_present_quarantine_history) == 0:
                        raise exceptions.ValidationException({'quarantine_room_id': messages.PRESENT_QUARANTINE_HISTORY_NOT_EXIST})
                    elif len(old_present_quarantine_history) >= 2:
                        raise exceptions.ValidationException({'quarantine_room_id': messages.MANY_PRESENT_QUARANTINE_HISTORY_EXIST})
                    else:
                        now_time = timezone.now()
                        old_present_quarantine_history = old_present_quarantine_history[0]
                        old_present_quarantine_history.status = QuarantineHistoryStatus.ENDED
                        old_present_quarantine_history.end_date = now_time
                        old_present_quarantine_history.end_type = QuarantineHistoryEndType.CHANGE_ROOM
                        old_present_quarantine_history.updated_by = request.user

                        new_quarantine_history = QuarantineHistory(
                            user=custom_user,
                            pandemic=custom_user.quarantine_ward.pandemic,
                            quarantine_ward=custom_user.quarantine_ward,
                            quarantine_room=member.quarantine_room,
                            status=QuarantineHistoryStatus.PRESENT,
                            start_date=now_time,
                            created_by=request.user,
                            updated_by=request.user,
                        )

                        new_quarantine_history.save()
                        old_present_quarantine_history.save()

            custom_user.save()
            member.save()

            self.do_after_change_room_of_member_work(member, old_room)

            member = Member.objects.get(id=member.id)

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            member_serializer = MemberSerializer(member, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['member'] = member_serializer.data

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='accept_one', detail=False)
    def accept_one_member(self, request):
        """Accept one member

        Args:
            + code: String
            - quarantine_room_id: int
            - quarantined_at: String vd:'2000-01-26T01:23:45.123456Z'
            - care_staff_code: String
        """

        accept_fields = [
            'code', 'quarantine_room_id',
            'quarantined_at', 'care_staff_code',
        ]

        require_fields = [
            'code',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['quarantined_at'])

            validator.extra_validate_to_accept_one_member()

            # check member
            custom_user = validator.get_field('custom_user')
            quarantine_room = validator.get_field('quarantine_room')
            quarantined_at = validator.get_field('quarantined_at')
            care_staff = validator.get_field('care_staff')

            if custom_user.role.name != 'MEMBER' or not hasattr(custom_user, 'member_x_custom_user'):
                raise exceptions.ValidationException({'main': messages.ISNOTMEMBER})
            if custom_user.status not in [CustomUserStatus.WAITING, CustomUserStatus.REFUSED]:
                raise exceptions.ValidationException({'main': messages.IS_NOT_WAITING_OR_REFUSED})

            must_not_empty_fields_of_custom_user = [
                'full_name', 'nationality', 'country', 'city', 'district',
                'ward', 'identity_number', 'quarantine_ward', 
            ]
            
            for field in must_not_empty_fields_of_custom_user:
                if not getattr(custom_user, field):
                    raise exceptions.ValidationException({field: messages.EMPTY})
            
            member = custom_user.member_x_custom_user

            # quarantined_at
            if not quarantined_at:
                quarantined_at = timezone.now()
            member.quarantined_at = quarantined_at

            if member.label == MemberLabel.F0:
                if custom_user.quarantine_ward.pandemic:
                    if member.number_of_vaccine_doses < 2:
                        remain_qt = custom_user.quarantine_ward.pandemic.remain_qt_pos_not_vac
                    else:
                        remain_qt = custom_user.quarantine_ward.pandemic.remain_qt_pos_vac
                else:
                    if member.number_of_vaccine_doses < 2:
                        remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                    else:
                        remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                member.quarantined_finish_expected_at = quarantined_at + datetime.timedelta(days=remain_qt)
            else:
                if custom_user.quarantine_ward.pandemic:
                    if member.number_of_vaccine_doses < 2:
                        remain_qt = custom_user.quarantine_ward.pandemic.quarantine_time_not_vac
                    else:
                        remain_qt = custom_user.quarantine_ward.pandemic.quarantine_time_vac
                else:
                    if member.number_of_vaccine_doses < 2:
                        remain_qt = int(os.environ.get('QUARANTINE_TIME_NOT_VAC', 14))
                    else:
                        remain_qt = int(os.environ.get('QUARANTINE_TIME_VAC', 10))
                member.quarantined_finish_expected_at = quarantined_at + datetime.timedelta(days=remain_qt)

            # room
            if not quarantine_room:
                # auto set room
                input_dict_for_get_suitable_room = dict()
                input_dict_for_get_suitable_room['quarantine_ward'] = custom_user.quarantine_ward
                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                input_dict_for_get_suitable_room['label'] = member.label
                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                input_dict_for_get_suitable_room['old_quarantine_room'] = None
                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                quarantine_room = suitable_room_dict['room']
                warning = suitable_room_dict['warning']
                if not quarantine_room:
                    raise exceptions.ValidationException({'main': warning})
                else:
                    member.quarantine_room = quarantine_room
            else:
                # check room received
                result = self.check_room_for_member(custom_user, quarantine_room)
                if result == messages.SUCCESS:
                    member.quarantine_room = quarantine_room
                else:
                    raise exceptions.ValidationException({'quarantine_room_id': result})

            # care_staff
            if not care_staff:
                if not member.care_staff:
                    # auto set care_staff
                    input_dict_for_get_care_staff = dict()
                    input_dict_for_get_care_staff['quarantine_floor'] = member.quarantine_room.quarantine_floor

                    suitable_care_staff_dict = self.get_suitable_care_staff_for_member(input_dict=input_dict_for_get_care_staff)
                    care_staff = suitable_care_staff_dict['care_staff']
                    warning = suitable_care_staff_dict['warning']
                    if care_staff:
                        member.care_staff = care_staff
            else:
                # check care_staff received
                if care_staff.quarantine_ward != custom_user.quarantine_ward:
                    raise exceptions.ValidationException({'care_staff_code': messages.NOT_IN_QUARANTINE_WARD_OF_MEMBER})
                member.care_staff = care_staff
            
            custom_user.status = CustomUserStatus.AVAILABLE
            custom_user.created_by = request.user
            custom_user.updated_by = request.user
            member.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()

            # create QuarantineHistory
            old_present_quarantine_history = QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT)
            if len(list(old_present_quarantine_history)) > 0:
                raise exceptions.ValidationException({'main': messages.PRESENT_QUARANTINE_HISTORY_EXIST})

            quarantine_history = QuarantineHistory(
                user=custom_user,
                pandemic=custom_user.quarantine_ward.pandemic,
                quarantine_ward=custom_user.quarantine_ward,
                quarantine_room=member.quarantine_room,
                status=QuarantineHistoryStatus.PRESENT,
                start_date=member.quarantined_at,
                created_by=request.user,
                updated_by=request.user,
            )

            quarantine_history.save()

            custom_user.save()
            member.save()

            # Send notification
            title = 'Xét duyệt tài khoản'
            description = f'Bạn đã được chấp nhận cách ly tại khu cách ly {custom_user.quarantine_ward.full_name}'
            create_and_send_noti_to_list_user(title=title, description=description, created_by=None, receive_user_list=[custom_user])

            self.do_after_change_room_of_member_work(member, None)

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='accept_many', detail=False)
    def accept_many_members(self, request):
        """Accept some members

        Args:
            + member_codes: String <code>,<code>
        """

        accept_fields = [
            'member_codes',
        ]

        require_fields = [
            'member_codes',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_accept_many_members()

            # accept members

            return_data = dict()

            custom_users = validator.get_field('members')

            for custom_user in custom_users:
                if custom_user.role.name != 'MEMBER' or not hasattr(custom_user, 'member_x_custom_user'):
                    return_data[custom_user.code] = messages.ISNOTMEMBER
                    continue

                if custom_user.status not in [CustomUserStatus.WAITING, CustomUserStatus.REFUSED]:
                    return_data[custom_user.code] = messages.IS_NOT_WAITING_OR_REFUSED
                    continue

                member = custom_user.member_x_custom_user

                must_not_empty_fields_of_custom_user = [
                    'full_name', 'nationality', 'country', 'city', 'district',
                    'ward', 'identity_number', 'quarantine_ward', 
                ]

                is_continue_another_custom_user = False
                for field in must_not_empty_fields_of_custom_user:
                    if not getattr(custom_user, field):
                        is_continue_another_custom_user = True
                        return_data[custom_user.code] = f'{field}: {messages.EMPTY}'
                        break
                if is_continue_another_custom_user:
                    continue
                
                # set room
                input_dict_for_get_suitable_room = dict()
                input_dict_for_get_suitable_room['quarantine_ward'] = custom_user.quarantine_ward
                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                input_dict_for_get_suitable_room['label'] = member.label
                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                input_dict_for_get_suitable_room['old_quarantine_room'] = None
                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                quarantine_room = suitable_room_dict['room']
                warning = suitable_room_dict['warning']
                if not quarantine_room:
                    return_data[custom_user.code] = warning
                    continue
                else:
                    member.quarantine_room = quarantine_room

                # set care_staff
                if not member.care_staff:
                    input_dict_for_get_care_staff = dict()
                    input_dict_for_get_care_staff['quarantine_floor'] = member.quarantine_room.quarantine_floor

                    suitable_care_staff_dict = self.get_suitable_care_staff_for_member(input_dict=input_dict_for_get_care_staff)
                    care_staff = suitable_care_staff_dict['care_staff']
                    warning = suitable_care_staff_dict['warning']
                    if care_staff:
                        member.care_staff = care_staff

                custom_user.status = CustomUserStatus.AVAILABLE

                # set quarantined_at
                quarantined_at = timezone.now()
                member.quarantined_at = quarantined_at
                if member.label == MemberLabel.F0:
                    if custom_user.quarantine_ward.pandemic:
                        if member.number_of_vaccine_doses < 2:
                            remain_qt = custom_user.quarantine_ward.pandemic.remain_qt_pos_not_vac
                        else:
                            remain_qt = custom_user.quarantine_ward.pandemic.remain_qt_pos_vac
                    else:
                        if member.number_of_vaccine_doses < 2:
                            remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                        else:
                            remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                    member.quarantined_finish_expected_at = quarantined_at + datetime.timedelta(days=remain_qt)
                else:
                    if custom_user.quarantine_ward.pandemic:
                        if member.number_of_vaccine_doses < 2:
                            remain_qt = custom_user.quarantine_ward.pandemic.quarantine_time_not_vac
                        else:
                            remain_qt = custom_user.quarantine_ward.pandemic.quarantine_time_vac
                    else:
                        if member.number_of_vaccine_doses < 2:
                            remain_qt = int(os.environ.get('QUARANTINE_TIME_NOT_VAC', 14))
                        else:
                            remain_qt = int(os.environ.get('QUARANTINE_TIME_VAC', 10))
                    member.quarantined_finish_expected_at = quarantined_at + datetime.timedelta(days=remain_qt)

                member.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()

                custom_user.created_by = request.user
                custom_user.updated_by = request.user

                # create QuarantineHistory
                old_present_quarantine_history = QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT)
                if len(list(old_present_quarantine_history)) > 0:
                    return_data[custom_user.code] = messages.PRESENT_QUARANTINE_HISTORY_EXIST
                    continue

                quarantine_history = QuarantineHistory(
                    user=custom_user,
                    pandemic=custom_user.quarantine_ward.pandemic,
                    quarantine_ward=custom_user.quarantine_ward,
                    quarantine_room=member.quarantine_room,
                    status=QuarantineHistoryStatus.PRESENT,
                    start_date=member.quarantined_at,
                    created_by=request.user,
                    updated_by=request.user,
                )

                quarantine_history.save()

                custom_user.save()
                member.save()

                # Send notification
                title = 'Xét duyệt tài khoản'
                description = f'Bạn đã được chấp nhận cách ly tại khu cách ly {custom_user.quarantine_ward.full_name}'
                create_and_send_noti_to_list_user(title=title, description=description, created_by=None, receive_user_list=[custom_user])

                self.do_after_change_room_of_member_work(member, None)

            return_message = messages.SUCCESS
            if return_data:
                return_message = messages.WARNING
            
            return self.response_handler.handle(data=return_data, message=return_message)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='refuse', detail=False)
    def refuse_members(self, request):
        """Refuse some members

        Args:
            + member_codes: String <code>,<code>
        """

        accept_fields = [
            'member_codes',
        ]

        require_fields = [
            'member_codes',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_refuse_member()

            # refuse members

            custom_users = validator.get_field('members')

            for custom_user in custom_users:
                if custom_user.role.name == 'MEMBER' and hasattr(custom_user, 'member_x_custom_user') and custom_user.status == CustomUserStatus.WAITING:
                    custom_user.status = CustomUserStatus.REFUSED
                    custom_user.updated_by = request.user
                    member = custom_user.member_x_custom_user
                    member.quarantine_room = None
                    custom_user.save()
                    member.save()

                    # Send notification
                    title = 'Xét duyệt tài khoản'
                    description = f'Bạn đã bị từ chối cách ly tại khu cách ly {custom_user.quarantine_ward.full_name}'
                    create_and_send_noti_to_list_user(title=title, description=description, created_by=None, receive_user_list=[custom_user])

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='finish_quarantine', detail=False)
    def finish_quarantine_members(self, request):
        """Finish quarantine some members

        Args:
            + member_codes: String <code>,<code>
        """

        accept_fields = [
            'member_codes',
        ]

        require_fields = [
            'member_codes',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_finish_quarantine_member()

            # finish quarantine members

            return_data = dict()

            custom_users = validator.get_field('members')

            for custom_user in custom_users:
                if custom_user.role.name != 'MEMBER' or not hasattr(custom_user, 'member_x_custom_user'):
                    return_data[custom_user.code] = messages.ISNOTMEMBER
                    continue

                if custom_user.status != CustomUserStatus.AVAILABLE:
                    return_data[custom_user.code] = messages.ISNOTAVAILABLE
                    continue

                if not validator.check_member_can_finish_quarantine(custom_user):
                    return_data[custom_user.code] = messages.CANNOT_FINISH_QUARANTINE
                    continue

                custom_user.status = CustomUserStatus.LEAVE
                custom_user.updated_by = request.user
                member = custom_user.member_x_custom_user
                member.quarantined_status = MemberQuarantinedStatus.COMPLETED
                member.quarantined_finished_at = timezone.now()

                old_room = member.quarantine_room
                member.quarantine_room = None

                # update QuarantineHistory
                old_present_quarantine_history = QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT)
                if len(old_present_quarantine_history) == 0:
                    return_data[custom_user.code] = messages.PRESENT_QUARANTINE_HISTORY_NOT_EXIST
                    continue
                elif len(old_present_quarantine_history) >= 2:
                    return_data[custom_user.code] = messages.MANY_PRESENT_QUARANTINE_HISTORY_EXIST
                    continue
                else:
                    old_present_quarantine_history = old_present_quarantine_history[0]
                    old_present_quarantine_history.status = QuarantineHistoryStatus.ENDED
                    old_present_quarantine_history.end_date = member.quarantined_finished_at
                    old_present_quarantine_history.end_type = QuarantineHistoryEndType.COMPLETED
                    old_present_quarantine_history.updated_by = request.user

                    old_present_quarantine_history.save()

                custom_user.save()
                member.save()

                self.do_after_change_room_of_member_work(member, old_room)
            
            return_message = messages.SUCCESS
            if return_data:
                return_message = messages.WARNING
            
            return self.response_handler.handle(data=return_data, message=return_message)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_member(self, request):
        """Get a list of members

        Args:
            - status_list: String <status>,<status> ['WAITING', 'REFUSED', 'LOCKED', 'AVAILABLE', 'LEAVE']
            - quarantined_status_list: String <status>,<status> ['COMPLETED', 'QUARANTINING', 'HOSPITALIZE_WAITING', 'HOSPITALIZE', 'MOVED']
            - health_status_list: String <status>,<status> ['NORMAL', 'UNWELL', 'SERIOUS', 'Null']
            - positive_test_now_list: String <value>,<value> ['True', 'False', 'Null']
            - is_last_tested: boolean - True để lọc những người cách ly đến hạn xét nghiệm, False hoặc không truyền đồng nghĩa không lọc
            - can_finish_quarantine: boolean - True để lọc những người cách ly có thể hoàn thành cách ly, False hoặc không truyền đồng nghĩa không lọc
            - is_need_change_room_because_be_positive: boolean
            - created_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - created_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_finish_expected_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_finish_expected_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_ward_id: String
            - quarantine_building_id: String
            - quarantine_floor_id: String
            - quarantine_room_id: String
            - label_list: String <label>,<label> ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            - care_staff_code: String
            - page: int
            - page_size: int
            - search: String
            - order_by: String ['quarantined_finished_at']
        """

        accept_fields = [
            'status_list', 'quarantined_status_list',
            'health_status_list', 'positive_test_now_list',
            'is_last_tested', 'can_finish_quarantine',
            'is_need_change_room_because_be_positive',
            'created_at_max', 'created_at_min',
            'quarantined_at_max', 'quarantined_at_min',
            'quarantined_finish_expected_at_max', 'quarantined_finish_expected_at_min',
            'quarantine_ward_id', 'quarantine_building_id',
            'quarantine_floor_id', 'quarantine_room_id',
            'label_list', 'care_staff_code',
            'page', 'page_size', 'search',
            'order_by',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)

            validator.is_valid_fields([
                'status_list', 'quarantined_status_list',
                'positive_test_now_list', 'health_status_list', 'is_last_tested',
                'can_finish_quarantine', 'is_need_change_room_because_be_positive',
                'created_at_max', 'created_at_min',
                'quarantined_at_max', 'quarantined_at_min',
                'quarantined_finish_expected_at_max', 'quarantined_finish_expected_at_min',
                'label_list', 'order_by',
            ])
            validator.extra_validate_to_filter_member()

            query_set = CustomUser.objects.all()

            list_to_filter_user = [key for key in accepted_fields.keys()]
            list_to_filter_user = set(list_to_filter_user) - \
            {'is_last_tested', 'is_need_change_room_because_be_positive', 'page', 'page_size'}
            list_to_filter_user = list(list_to_filter_user) + \
            [
                'status_list',
                'last_tested_max', 'role_name', 'quarantined_at_max',
                'quarantined_finish_expected_at_max',
                'positive_test_now_list', 'health_status_list',
            ]

            dict_to_filter_user = validator.get_data(list_to_filter_user)

            # Check ward of sender
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if hasattr(validator, '_quarantine_ward'):
                    # Sender want filter with ward, building, floor or room
                    if validator.get_field('quarantine_ward') != request.user.quarantine_ward:
                        raise exceptions.AuthenticationException({'quarantine_ward_id': messages.NO_PERMISSION})
                else:
                    dict_to_filter_user['quarantine_ward_id'] = request.user.quarantine_ward.id

            if 'order_by' not in accepted_fields.keys():
                dict_to_filter_user.setdefault('order_by', '-quarantined_at')

            filter = MemberFilter(dict_to_filter_user, queryset=query_set)

            query_set = filter.qs

            query_set = query_set.select_related('member_x_custom_user__quarantine_room__quarantine_floor__quarantine_building', 'quarantine_ward')

            serializer = FilterMemberSerializer(query_set, many=True)

            # check if filter member is need change room because be positive
            is_need_change_room_because_be_positive = validator.get_field('is_need_change_room_because_be_positive')
            if is_need_change_room_because_be_positive == True:
                # filter user that positive_test_now = True and need change room
                result_users = list(query_set)
                remain_result_users = [user for user in result_users if self.count_positive_test_now_not_true_in_room(user.member_x_custom_user.quarantine_room) >= 1]
                serializer = FilterMemberSerializer(remain_result_users, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='not_member_filter', detail=False)
    def not_member_filter(self, request):
        """Get a list user of some role

        Args:
            + role_name_list: String <role_name>,<role_name> ['MEMBER', 'SUPER_MANAGER', 'MANAGER', 'ADMINISTRATOR', 'STAFF']
            - quarantine_ward_id: String
        """

        
        accept_fields = [
            'role_name_list', 'quarantine_ward_id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(['role_name_list',])
            validator.is_valid_fields([
                'role_name_list',
            ])
            validator.extra_validate_to_filter_not_member()

            dict_to_filter_user = validator.get_data(['role_name_list', 'quarantine_ward_id', 'status'])

            # Check ward of sender
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if hasattr(validator, '_quarantine_ward'):
                    # Sender want filter with ward, building, floor or room
                    if validator.get_field('quarantine_ward') != request.user.quarantine_ward:
                        raise exceptions.AuthenticationException({'quarantine_ward_id': messages.NO_PERMISSION})
                else:
                    dict_to_filter_user['quarantine_ward_id'] = request.user.quarantine_ward.id

            dict_to_filter_user.setdefault('order_by', '-created_at')

            query_set = CustomUser.objects.all()

            filter = UserFilter(dict_to_filter_user, queryset=query_set)

            query_set = filter.qs

            if 'SUPER_MANAGER' in validator.get_field('new_role_name_list'):
                super_manager_query_set = CustomUser.objects.filter(role__name='SUPER_MANAGER', status=CustomUserStatus.AVAILABLE)
                query_set = (query_set | super_manager_query_set).distinct()

            query_set = query_set.select_related('role')

            serializer = FilterNotMemberSerializer(query_set, many=True)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get_suitable_room', detail=False)
    def get_suitable_room(self, request):
        """Get a suitable room for a member

        Args:
            + quarantine_ward_id: int
            + gender: String ['MALE', 'FEMALE']
            + label: String ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            + number_of_vaccine_doses: int
            + positive_test_now: String ['True', 'False', 'Null']
            - old_quarantine_room_id: int
            - not_quarantine_room_ids: String <id>,<id>,<id>
        """

        
        accept_fields = [
            'quarantine_ward_id', 'gender',
            'label', 'number_of_vaccine_doses',
            'positive_test_now',
            'old_quarantine_room_id', 'not_quarantine_room_ids',
        ]

        require_fields = [
            'quarantine_ward_id', 'gender',
            'label', 'number_of_vaccine_doses',
            'positive_test_now',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'gender', 'label', 'number_of_vaccine_doses',
                'positive_test_now', 'not_quarantine_room_ids',
            ])

            validator.extra_validate_to_get_suitable_room()

            input_dict_for_get_suitable_room = validator.get_data([
                'quarantine_ward', 'gender',
                'label', 'number_of_vaccine_doses',
                'positive_test_now',
                'old_quarantine_room', 'not_quarantine_room_ids',
            ])

            return_value = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
            room = return_value['room']
            warning = return_value['warning']

            return_dict = dict()
            return_dict['quarantine_room'] = None

            if room:
                return_message = messages.SUCCESS
                error_code = 0
                serializer = QuarantineRoomSerializer(room, many=False)
                return_dict['quarantine_room'] = serializer.data
                floor = room.quarantine_floor
                serializer = QuarantineFloorSerializer(floor, many=False)
                return_dict['quarantine_floor'] = serializer.data
                building = floor.quarantine_building
                serializer = QuarantineBuildingSerializer(building, many=False)
                return_dict['quarantine_building'] = serializer.data
                ward = building.quarantine_ward
                serializer = BaseQuarantineWardSerializer(ward, many=False)
                return_dict['quarantine_ward'] = serializer.data
            else:
                return_message = warning
                error_code = 400

            return self.response_handler.handle(error_code=error_code, message=return_message, data=return_dict)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='change_quarantine_ward_and_room', detail=False)
    def change_quarantine_ward_and_room(self, request):
        """Change quarantine_ward and room of an available member, can change care staff too

        Args:
            + custom_user_code: String
            + quarantine_ward_id: int
            - quarantine_room_id: int
            - care_staff_code: String
        """

        accept_fields = [
            'custom_user_code', 'quarantine_ward_id',
            'quarantine_room_id', 'care_staff_code',
        ]

        require_fields = [
            'custom_user_code', 'quarantine_ward_id',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_change_quarantine_ward_and_room_of_available_member()

            # check member
            custom_user = validator.get_field('custom_user')
            quarantine_ward = validator.get_field('quarantine_ward')
            quarantine_room = validator.get_field('quarantine_room')
            care_staff = validator.get_field('care_staff')

            if custom_user.role.name != 'MEMBER' or not hasattr(custom_user, 'member_x_custom_user'):
                raise exceptions.ValidationException({'main': messages.ISNOTMEMBER})
            if custom_user.status != CustomUserStatus.AVAILABLE:
                raise exceptions.ValidationException({'main': messages.ISNOTAVAILABLE})
            
            member = custom_user.member_x_custom_user

            custom_user.quarantine_ward = quarantine_ward

            old_room = member.quarantine_room

            # room
            if not quarantine_room:
                # auto set room
                input_dict_for_get_suitable_room = dict()
                input_dict_for_get_suitable_room['quarantine_ward'] = quarantine_ward
                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                input_dict_for_get_suitable_room['label'] = member.label
                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                input_dict_for_get_suitable_room['old_quarantine_room'] = member.quarantine_room
                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                quarantine_room = suitable_room_dict['room']
                warning = suitable_room_dict['warning']
                if not quarantine_room:
                    raise exceptions.ValidationException({'main': warning})
                else:
                    new_room = quarantine_room
            else:
                # check room received
                result = self.check_room_for_member(custom_user, quarantine_room)
                if result == messages.SUCCESS:
                    new_room = quarantine_room
                else:
                    raise exceptions.ValidationException({'quarantine_room_id': result})
            if new_room != old_room:
                member.quarantine_room = new_room

            # care_staff
            if not care_staff:
                if not member.care_staff or member.care_staff.quarantine_ward != custom_user.quarantine_ward:
                    # auto set care_staff
                    input_dict_for_get_care_staff = dict()
                    input_dict_for_get_care_staff['quarantine_floor'] = member.quarantine_room.quarantine_floor

                    suitable_care_staff_dict = self.get_suitable_care_staff_for_member(input_dict=input_dict_for_get_care_staff)
                    care_staff = suitable_care_staff_dict['care_staff']

                    member.care_staff = care_staff
            else:
                # check care_staff received
                if care_staff.quarantine_ward != custom_user.quarantine_ward:
                    raise exceptions.ValidationException({'care_staff_code': messages.NOT_IN_QUARANTINE_WARD_OF_MEMBER})
                member.care_staff = care_staff
            
            custom_user.updated_by = request.user

            if new_room != old_room:
                # update QuarantineHistory
                old_present_quarantine_history = list(QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT))
                if len(old_present_quarantine_history) == 0:
                    raise exceptions.ValidationException({'main': messages.PRESENT_QUARANTINE_HISTORY_NOT_EXIST})
                elif len(old_present_quarantine_history) >= 2:
                    raise exceptions.ValidationException({'main': messages.MANY_PRESENT_QUARANTINE_HISTORY_EXIST})
                else:
                    now_time = timezone.now()
                    old_present_quarantine_history = old_present_quarantine_history[0]
                    old_present_quarantine_history.status = QuarantineHistoryStatus.ENDED
                    old_present_quarantine_history.end_date = now_time
                    old_present_quarantine_history.end_type = QuarantineHistoryEndType.CHANGE_ROOM
                    old_present_quarantine_history.updated_by = request.user

                    new_quarantine_history = QuarantineHistory(
                        user=custom_user,
                        pandemic=custom_user.quarantine_ward.pandemic,
                        quarantine_ward=custom_user.quarantine_ward,
                        quarantine_room=member.quarantine_room,
                        status=QuarantineHistoryStatus.PRESENT,
                        start_date=now_time,
                        created_by=request.user,
                        updated_by=request.user,
                    )

                    new_quarantine_history.save()
                    old_present_quarantine_history.save()

            custom_user.save()
            member.save()

            if new_room != old_room:
                self.do_after_change_room_of_member_work(member, old_room)

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='hospitalize', detail=False)
    def hospitalize(self, request):
        """Hospitalize a 'AVAILABLE' member.

        Args:
            + code: String
            + hospital_name: String
            - note: String
        """

        
        accept_fields = [
            'code', 'hospital_name', 'note'
        ]

        require_fields = [
            'code', 'hospital_name'
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_hospitalize()

            custom_user = validator.get_field('custom_user')
            member = custom_user.member_x_custom_user
            hospital_name = validator.get_field('hospital_name')
            note = validator.get_field('note')

            if hospital_name == 'Bệnh viện dã chiến':
                # call api nhom kia
                benhvien_url = 'https://api.bvdc.link/api/transfer'
                data = {
                    "cmnd": "",
                    "name": "",
                    "phone": "",
                    "birthDay": "",
                    "email": "",
                    "gender": "",
                    "tinh": "",
                    "huyen": "",
                    "xa": "",
                    "thon": "",
                    "bhyt": "",
                    "sickness": [],
                }
                data['cmnd'] = custom_user.identity_number if custom_user.identity_number else ''
                data['name'] = custom_user.full_name if custom_user.full_name else ''
                data['phone'] = custom_user.phone_number if custom_user.phone_number else ''

                if custom_user.birthday:
                    data['birthDay'] = custom_user.birthday[6:10] + '-' + custom_user.birthday[3:5] + '-' + custom_user.birthday[0:2]
                else:
                    data['birthDay'] = ''

                data['email'] = custom_user.email if custom_user.email else ''
                data['gender'] = 'NAM' if custom_user.gender == Gender.MALE else 'NU'
                data['tinh'] = custom_user.city.name if custom_user.city else ''
                data['huyen'] = custom_user.district.name if custom_user.district else ''
                data['xa'] = custom_user.ward.name if custom_user.ward else ''
                data['thon'] = custom_user.detail_address if custom_user.detail_address else ''
                data['bhyt'] = custom_user.health_insurance_number if custom_user.health_insurance_number else ''
                background_disease_names = []
                if custom_user.member_x_custom_user.background_disease:
                    background_disease_ids = split_input_list(custom_user.member_x_custom_user.background_disease)
                    background_disease_objects = [self.get_background_disease_by_id(id) for id in background_disease_ids]
                    background_disease_names = [item.name for item in background_disease_objects if item]
                data['sickness'] = background_disease_names

                data = json.dumps(data)

                headers = {'Content-type': 'application/json; charset=UTF-8'}

                r = requests.post(benhvien_url, headers=headers, data=data)

                benhvien_response = r.json()
                if benhvien_response['success'] == True:
                    # Success register hospitalize, but need wait for accept
                    member.quarantined_status = MemberQuarantinedStatus.HOSPITALIZE_WAITING
                    member.save()
                else:
                    raise exceptions.ValidationException({'main': messages.CANNOT_HOSPITALIZE_THIS_HOSPITAL})
            else:
                # hospitalize
                custom_user.status = CustomUserStatus.LEAVE
                member.quarantined_status = MemberQuarantinedStatus.HOSPITALIZE
                member.quarantined_finished_at = timezone.now()

                old_room = member.quarantine_room
                member.quarantine_room = None

                # update QuarantineHistory
                old_present_quarantine_history = QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT)
                if len(old_present_quarantine_history) == 0:
                    raise exceptions.ValidationException({'main': messages.PRESENT_QUARANTINE_HISTORY_NOT_EXIST})
                elif len(old_present_quarantine_history) >= 2:
                    raise exceptions.ValidationException({'main': messages.MANY_PRESENT_QUARANTINE_HISTORY_EXIST})
                else:
                    old_present_quarantine_history = old_present_quarantine_history[0]
                    old_present_quarantine_history.status = QuarantineHistoryStatus.ENDED
                    old_present_quarantine_history.end_date = member.quarantined_finished_at
                    old_present_quarantine_history.end_type = QuarantineHistoryEndType.HOSPITALIZE
                    if old_present_quarantine_history.note:
                        old_present_quarantine_history.note += ';' + hospital_name
                        old_present_quarantine_history.note += (';' + note) if note else ''
                    else:
                        old_present_quarantine_history.note = hospital_name
                        old_present_quarantine_history.note += (';' + note) if note else ''
                    old_present_quarantine_history.updated_by = request.user

                old_present_quarantine_history.save()
                custom_user.save()
                member.save()

                self.do_after_change_room_of_member_work(member, old_room)

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='hospitalize_confirm', detail=False)
    def bvdc_call_hospitalize_confirm(self, request):
        """For benh vien da chien to confirm hospitalize for a hospitalize waiting member

        Headers:
            + secret_key: String
        Args:
            + phone_number: String
            + confirm: ['ACCEPT', 'REFUSE']
        """

        
        accept_fields = [
            'phone_number', 'confirm',
        ]

        require_fields = [
            'phone_number', 'confirm',
        ]

        try:
            if request.headers.get('secret-key') != os.environ.get('BENH_VIEN_DA_CHIEN_SECRET_KEY', 'K6YNR4C6UVQMOGPB'):
                raise exceptions.ValidationException({'secret-key': messages.INVALID})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)

            validator.extra_validate_to_bvdc_call_hospitalize_confirm()

            custom_user = validator.get_field('custom_user')
            member = custom_user.member_x_custom_user
            confirm = validator.get_field('confirm')
            
            if confirm == 'ACCEPT':
                # hospitalize accept
                custom_user.status = CustomUserStatus.LEAVE
                member.quarantined_status = MemberQuarantinedStatus.HOSPITALIZE
                member.quarantined_finished_at = timezone.now()

                old_room = member.quarantine_room
                member.quarantine_room = None

                # update QuarantineHistory
                try:
                    old_present_quarantine_history = QuarantineHistory.objects.filter(user=custom_user, status=QuarantineHistoryStatus.PRESENT)
                    if len(old_present_quarantine_history) == 0:
                        raise exceptions.ValidationException({'main': messages.PRESENT_QUARANTINE_HISTORY_NOT_EXIST})
                    elif len(old_present_quarantine_history) >= 2:
                        raise exceptions.ValidationException({'main': messages.MANY_PRESENT_QUARANTINE_HISTORY_EXIST})
                    else:
                        old_present_quarantine_history = old_present_quarantine_history[0]
                        old_present_quarantine_history.status = QuarantineHistoryStatus.ENDED
                        old_present_quarantine_history.end_date = member.quarantined_finished_at
                        old_present_quarantine_history.end_type = QuarantineHistoryEndType.HOSPITALIZE
                        if old_present_quarantine_history.note:
                            old_present_quarantine_history.note += ';' + 'Bệnh viện dã chiến'
                        else:
                            old_present_quarantine_history.note = 'Bệnh viện dã chiến'
                    old_present_quarantine_history.save()
                except:
                    ...

                custom_user.save()
                member.save()

                try:
                    self.do_after_change_room_of_member_work(member, old_room)
                except:
                    ...
            else:
                # hospitalize refuse
                member.quarantined_status = MemberQuarantinedStatus.QUARANTINING
                member.save()

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='member_call_requarantine', detail=False)
    def member_call_requarantine(self, request):
        """Requarantine a 'LEAVE' member. This member will be added to waiting member list. This member call api for themselve.

        Args:
            + quarantine_ward_id: int
            + label: String ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            - quarantine_reason: String
            - positive_tested_before: boolean
            - first_positive_test_date: String vd:'2000-01-26T01:23:45.123456Z'
        """

        
        accept_fields = [
            'quarantine_ward_id', 'label',
            'quarantine_reason', 'positive_tested_before',
            'first_positive_test_date',
        ]

        require_fields = [
            'quarantine_ward_id', 'label',
        ]

        try:
            custom_user = request.user
            if custom_user.role.name != 'MEMBER' or not hasattr(custom_user, 'member_x_custom_user'):
                raise exceptions.ValidationException({'sender': messages.ISNOTMEMBER})
            if custom_user.status != CustomUserStatus.LEAVE:
                raise exceptions.ValidationException({'sender': messages.ISNOTLEAVE})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['label', 'positive_tested_before', 'first_positive_test_date'])

            validator.extra_validate_to_member_call_requarantine(request.user)

            member = custom_user.member_x_custom_user

            quarantine_ward = validator.get_field('quarantine_ward')
            label = validator.get_field('label')
            quarantine_reason = validator.get_field('quarantine_reason')
            positive_tested_before = validator.get_field('positive_tested_before')
            first_positive_test_date = validator.get_field('first_positive_test_date')

            # requarantine
            custom_user.status = CustomUserStatus.WAITING
            custom_user.quarantine_ward = quarantine_ward
            member.quarantined_at = None
            member.quarantined_finish_expected_at = None
            member.quarantined_finished_at = None
            member.quarantined_status = MemberQuarantinedStatus.QUARANTINING
            if label == MemberLabel.F0:
                member.positive_test_now = True
            else:
                member.positive_test_now = None
            member.care_staff = None
            member.quarantine_room = None
            member.label = label
            member.quarantine_reason = quarantine_reason
            if positive_tested_before != None:
                member.positive_tested_before = positive_tested_before
            if 'first_positive_test_date' in accepted_fields.keys():
                member.first_positive_test_date = first_positive_test_date
            
            custom_user.save()
            member.save()

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='manager_call_requarantine', detail=False)
    def manager_call_requarantine(self, request):
        """Requarantine a 'LEAVE' member. This member will be AVAILABLE and have a room. Manager, Super Manager, Administrator can call this api.

        Args:
            + code: String
            + quarantine_ward_id: int
            + label: String ['F0', 'F1', 'F2', 'F3', 'FROM_EPIDEMIC_AREA', 'ABROAD']
            - quarantine_reason: String
            - positive_tested_before: boolean
            - first_positive_test_date: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantined_at: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_room_id: int
            - care_staff_code: String
        """

        accept_fields = [
            'code',
            'quarantine_ward_id', 'label', 'positive_tested_before',
            'first_positive_test_date',
            'quarantined_at', 'quarantine_reason',
            'quarantine_room_id', 'care_staff_code',
        ]

        require_fields = [
            'code', 'quarantine_ward_id', 'label',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['label', 'positive_tested_before', 'first_positive_test_date', 'quarantined_at'])
            
            validator.extra_validate_to_manager_call_requarantine()

            # update CustomUser
            custom_user = validator.get_field('custom_user')
            quarantine_ward = validator.get_field('quarantine_ward')
            custom_user.status = CustomUserStatus.AVAILABLE
            custom_user.quarantine_ward = quarantine_ward
            custom_user.updated_by = request.user

            # update Member
            member = custom_user.member_x_custom_user

            list_to_update_member = [key for key in accepted_fields.keys()]
            list_to_update_member = set(list_to_update_member) - \
            {'code', 'quarantine_ward_id', 'quarantine_room_id', 'care_staff_code',}
            list_to_update_member = list(list_to_update_member) + [
                'positive_test_now', 'quarantined_finish_expected_at',
                'quarantined_at', 'quarantine_reason',
                'quarantined_finished_at', 'quarantined_status',
            ]

            dict_to_update_member = validator.get_data(list_to_update_member)

            for attr, value in dict_to_update_member.items(): 
                setattr(member, attr, value)
            
            # extra set room for this member
            if hasattr(validator, '_quarantine_room'):
                # this field is received and not None
                quarantine_room = validator.get_field('quarantine_room')
                check_room_result = self.check_room_for_member(custom_user, quarantine_room)
                if check_room_result != messages.SUCCESS:
                    raise exceptions.ValidationException({'quarantine_room_id': check_room_result})
            else:
                input_dict_for_get_suitable_room = dict()
                input_dict_for_get_suitable_room['quarantine_ward'] = custom_user.quarantine_ward
                input_dict_for_get_suitable_room['gender'] = custom_user.gender
                input_dict_for_get_suitable_room['label'] = member.label
                input_dict_for_get_suitable_room['positive_test_now'] = member.positive_test_now
                input_dict_for_get_suitable_room['number_of_vaccine_doses'] = member.number_of_vaccine_doses
                input_dict_for_get_suitable_room['old_quarantine_room'] = None
                input_dict_for_get_suitable_room['not_quarantine_room_ids'] = []

                suitable_room_dict = self.get_suitable_room_for_member(input_dict=input_dict_for_get_suitable_room)
                quarantine_room = suitable_room_dict['room']
                warning = suitable_room_dict['warning']
                if not quarantine_room:
                    raise exceptions.ValidationException({'main': warning})

            member.quarantine_room = quarantine_room

            # care_staff
            if hasattr(validator, '_care_staff'):
                # care_staff received
                member.care_staff = validator.get_field('care_staff')
            else:
                if not member.care_staff or member.care_staff.quarantine_ward != custom_user.quarantine_ward:
                    # auto set care_staff
                    input_dict_for_get_care_staff = dict()
                    input_dict_for_get_care_staff['quarantine_floor'] = member.quarantine_room.quarantine_floor

                    suitable_care_staff_dict = self.get_suitable_care_staff_for_member(input_dict=input_dict_for_get_care_staff)
                    care_staff = suitable_care_staff_dict['care_staff']

                    member.care_staff = care_staff

            # create QuarantineHistory
            quarantine_history = QuarantineHistory(
                user=custom_user,
                pandemic=custom_user.quarantine_ward.pandemic,
                quarantine_ward=custom_user.quarantine_ward,
                quarantine_room=member.quarantine_room,
                status=QuarantineHistoryStatus.PRESENT,
                start_date=member.quarantined_at,
                created_by=request.user,
                updated_by=request.user,
            )

            custom_user.save()
            member.save()

            quarantine_history.save()

            self.do_after_change_room_of_member_work(member, None)

            member = Member.objects.get(id=member.id)

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            member_serializer = MemberSerializer(member, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['member'] = member_serializer.data

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class ManagerAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_manager(self, request):
        """Create a manager

        Args:
            + full_name: String
            + phone_number: String
            - email: String
            + birthday: String 'dd/mm/yyyy'
            + gender: String ['MALE', 'FEMALE']
            + nationality_code: String
            + country_code: int
            + city_id: int
            + district_id: int
            + ward_id: int
            + detail_address: String
            - professional: String in Professional enum, call api to get list of professionals
            - health_insurance_number: String
            + identity_number: String
            - passport_number: String
            + quarantine_ward_id: int
        """

        accept_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'professional',
            'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        require_fields = [
            'full_name', 'phone_number',
            'birthday', 'gender',
            'identity_number', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'quarantine_ward_id',
        ]

        custom_user_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'professional',
            'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        manager_fields = []

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'phone_number', 'email', 'birthday', 'gender',
                'professional',
                'passport_number', 'health_insurance_number', 'identity_number',
            ])
            validator.extra_validate_to_create_manager()

            # create CustomUser

            list_to_create_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_create_custom_user = set(list_to_create_custom_user) - \
            {'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_create_custom_user = list(list_to_create_custom_user) + \
            ['nationality', 'country', 'city', 'district', 'ward', 'quarantine_ward']

            dict_to_create_custom_user = validator.get_data(list_to_create_custom_user)

            custom_user = CustomUser(**dict_to_create_custom_user)
            custom_user.set_password('123456')
            custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            while (validator.is_code_exist(custom_user.code)):
                custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            custom_user.created_by = request.user
            custom_user.updated_by = request.user
            custom_user.role = Role.objects.get(name='MANAGER')

            # create Manager

            manager = Manager()
            manager.custom_user = custom_user

            custom_user.save()
            manager.save()

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            manager_serializer = ManagerSerializer(manager, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['manager'] = manager_serializer.data
            
            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_manager(self, request):
        """Update a manager, if dont get code, will update manager sending request

        Args:
            - code: int
            - full_name: String
            - email: String
            - birthday: String 'dd/mm/yyyy'
            - gender: String ['MALE', 'FEMALE']
            - professional: String in Professional enum, call api to get list of professionals
            - nationality_code: String
            - country_code: int
            - city_id: int
            - district_id: int
            - ward_id: int
            - detail_address: String
            - health_insurance_number: String
            - identity_number: String
            - passport_number: String
            - quarantine_ward_id: int
        """

        accept_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender', 'professional',
            'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        custom_user_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender',
            'professional', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        manager_fields = []

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            if 'code' not in accepted_fields.keys():
                accepted_fields['code'] = request.user.code

            validator = UserValidator(**accepted_fields)
            validator.is_valid_fields([
                'email', 'birthday', 'gender',
                'professional', 'passport_number',
                'health_insurance_number', 'identity_number',
            ])
            validator.extra_validate_to_update_manager()

            # update CustomUser

            custom_user = validator.get_field('custom_user')

            if request.user.role.name == 'MANAGER' and request.user != custom_user:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            list_to_update_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_update_custom_user = set(list_to_update_custom_user) - \
            {'code', 'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_update_custom_user = list(list_to_update_custom_user) + \
            ['nationality', 'country', 'city', 'district', 'ward', 'quarantine_ward']
            dict_to_update_custom_user = validator.get_data(list_to_update_custom_user)

            for attr, value in dict_to_update_custom_user.items(): 
                setattr(custom_user, attr, value)

            custom_user.updated_by = request.user

            response_data = dict()
            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            response_data['custom_user'] = custom_user_serializer.data

            custom_user.save()

            if hasattr(custom_user, 'manager_x_custom_user') and custom_user.manager_x_custom_user:
                manager = custom_user.manager_x_custom_user
                manager_serializer = ManagerSerializer(manager, many=False)
                response_data['manager'] = manager_serializer.data

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_manager(self, request):
        """Get a list of manager and super managers

        Args:
            - status_list: String ['WAITING', 'REFUSED', 'LOCKED', 'AVAILABLE', 'LEAVE']
            - health_status_list: String <status>,<status> ['NORMAL', 'UNWELL', 'SERIOUS', 'Null']
            - positive_test_now_list: String <value>,<value> ['True', 'False', 'Null']
            - is_last_tested: boolean - True để lọc những người quản lý đến hạn xét nghiệm, False hoặc không truyền đồng nghĩa không lọc
            - created_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - created_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_ward_id: int
            - page: int
            - page_size: int
            - search: String
            - order_by: String ['full_name', 'created_at']
        """

        accept_fields = [
            'status_list', 'health_status_list', 'positive_test_now_list',
            'is_last_tested',
            'created_at_max', 'created_at_min',
            'quarantine_ward_id',
            'page', 'page_size', 'search', 'order_by',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)

            validator.is_valid_fields([
                'status_list', 'health_status_list', 'positive_test_now_list',
                'is_last_tested',
                'created_at_max', 'created_at_min',
            ])
            validator.extra_validate_to_filter_manager()

            query_set = CustomUser.objects.all()

            list_to_filter_manager = [key for key in accepted_fields.keys()]
            list_to_filter_manager = set(list_to_filter_manager) - \
            {'is_last_tested', 'page', 'page_size'}
            list_to_filter_manager = list(list_to_filter_manager) + \
            [
                'status_list',
                'last_tested_max', 'role_name_list',
            ]

            dict_to_filter_manager = validator.get_data(list_to_filter_manager)

            # Check ward of sender
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if hasattr(validator, '_quarantine_ward'):
                    # Sender want filter with ward, building, floor or room
                    if validator.get_field('quarantine_ward') != request.user.quarantine_ward:
                        raise exceptions.AuthenticationException({'quarantine_ward_id': messages.NO_PERMISSION})
                else:
                    dict_to_filter_manager['quarantine_ward_id'] = request.user.quarantine_ward.id
            
            filter = ManagerFilter(dict_to_filter_manager, queryset=query_set)

            query_set = filter.qs

            super_manager_query_set = CustomUser.objects.filter(role__name='SUPER_MANAGER', status=CustomUserStatus.AVAILABLE)
            query_set = (query_set | super_manager_query_set).distinct()

            query_set = query_set.select_related('quarantine_ward', 'role')

            serializer = FilterManagerSerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class StaffAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_staff(self, request):
        """Create a staff

        Args:
            + full_name: String
            + phone_number: String
            - email: String
            + birthday: String 'dd/mm/yyyy'
            + gender: String ['MALE', 'FEMALE']
            + nationality_code: String
            + country_code: int
            + city_id: int
            + district_id: int
            + ward_id: int
            + detail_address: String
            - professional: String in Professional enum, call api to get list of professionals
            - health_insurance_number: String
            + identity_number: String
            - passport_number: String
            + quarantine_ward_id: int
            - care_area: String <id>,<id>,<id> trong đó <id> là id của tầng (QuarantineFloor)
        """

        accept_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'professional',
            'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id', 'care_area',
        ]

        require_fields = [
            'full_name', 'phone_number',
            'birthday', 'gender',
            'identity_number', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'quarantine_ward_id',
        ]

        custom_user_fields = [
            'full_name', 'phone_number', 'email',
            'birthday', 'gender', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'professional',
            'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'phone_number', 'email', 'birthday', 'gender',
                'passport_number', 'health_insurance_number', 'identity_number',
                'care_area', 'professional',
            ])
            validator.extra_validate_to_create_staff()

            # create CustomUser

            list_to_create_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_create_custom_user = set(list_to_create_custom_user) - \
            {'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_create_custom_user = list(list_to_create_custom_user) + \
            ['nationality', 'country', 'city', 'district', 'ward', 'quarantine_ward']

            dict_to_create_custom_user = validator.get_data(list_to_create_custom_user)

            custom_user = CustomUser(**dict_to_create_custom_user)
            custom_user.set_password('123456')
            custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            while (validator.is_code_exist(custom_user.code)):
                custom_user.code = custom_user_code_generator(custom_user.quarantine_ward.id)
            custom_user.created_by = request.user
            custom_user.updated_by = request.user
            custom_user.role = Role.objects.get(name='STAFF')

            # create Staff

            dict_to_create_staff = validator.get_data(['care_area'])
            staff = Staff(**dict_to_create_staff)
            staff.custom_user = custom_user

            custom_user.save()
            staff.save()

            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            staff_serializer = StaffSerializer(staff, many=False)
            
            response_data = dict()
            response_data['custom_user'] = custom_user_serializer.data
            response_data['staff'] = staff_serializer.data
            
            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_staff(self, request):
        """Update a staff, if dont get code, will update staff sending request

        Args:
            - code: int
            - full_name: String
            - email: String
            - birthday: String 'dd/mm/yyyy'
            - gender: String ['MALE', 'FEMALE']
            - professional: String in Professional enum, call api to get list of professionals
            - nationality_code: String
            - country_code: int
            - city_id: int
            - district_id: int
            - ward_id: int
            - detail_address: String
            - health_insurance_number: String
            - identity_number: String
            - passport_number: String
            - quarantine_ward_id: int
            - care_area: String <id>,<id>,<id> trong đó <id> là id của tầng (QuarantineFloor)
        """

        accept_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender',
            'professional', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id', 'care_area',
        ]

        custom_user_fields = [
            'code', 'full_name', 'email',
            'birthday', 'gender',
            'professional', 'nationality_code',
            'country_code', 'city_id', 'district_id', 'ward_id',
            'detail_address', 'health_insurance_number',
            'identity_number', 'passport_number',
            'quarantine_ward_id',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            if 'code' not in accepted_fields.keys():
                accepted_fields['code'] = request.user.code

            validator = UserValidator(**accepted_fields)
            validator.is_valid_fields([
                'email', 'birthday', 'gender',
                'professional', 'passport_number',
                'health_insurance_number', 'identity_number',
                'care_area',
            ])
            validator.extra_validate_to_update_staff()

            # update CustomUser

            custom_user = validator.get_field('custom_user')

            if request.user.role.name == 'STAFF' and request.user != custom_user:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            list_to_update_custom_user = [key for key in accepted_fields.keys() if key in custom_user_fields]
            list_to_update_custom_user = set(list_to_update_custom_user) - \
            {'code', 'nationality_code', 'country_code', 'city_id', 'district_id', 'ward_id', 'quarantine_ward_id'}
            list_to_update_custom_user = list(list_to_update_custom_user) + \
            ['nationality', 'country', 'city', 'district', 'ward', 'quarantine_ward']
            dict_to_update_custom_user = validator.get_data(list_to_update_custom_user)

            for attr, value in dict_to_update_custom_user.items(): 
                setattr(custom_user, attr, value)

            custom_user.updated_by = request.user

            response_data = dict()
            custom_user_serializer = CustomUserSerializer(custom_user, many=False)
            response_data['custom_user'] = custom_user_serializer.data

            if hasattr(custom_user, 'staff_x_custom_user') and custom_user.staff_x_custom_user:
                staff = custom_user.staff_x_custom_user
                dict_to_update_staff = validator.get_data(['care_area'])

                for attr, value in dict_to_update_staff.items(): 
                    setattr(staff, attr, value)

                staff_serializer = StaffSerializer(staff, many=False)
                response_data['staff'] = staff_serializer.data

                staff.save()

            custom_user.save()  

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_staff(self, request):
        """Get a list of staffs

        Args:
            - status_list: String ['WAITING', 'REFUSED', 'LOCKED', 'AVAILABLE', 'LEAVE']
            - health_status_list: String <status>,<status> ['NORMAL', 'UNWELL', 'SERIOUS', 'Null']
            - positive_test_now_list: String <value>,<value> ['True', 'False', 'Null']
            - is_last_tested: boolean - True để lọc những người cán bộ đến hạn xét nghiệm, False hoặc không truyền đồng nghĩa không lọc
            - created_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - created_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_ward_id: int
            - care_area: String <id>,<id> (id của tầng)
            - page: int
            - page_size: int
            - search: String
            - order_by: String ['full_name', 'created_at'], mặc định sắp thứ tự theo số lượng người cách ly đang chăm sóc tăng dần
        """

        accept_fields = [
            'status_list', 'health_status_list', 'positive_test_now_list',
            'is_last_tested',
            'created_at_max', 'created_at_min',
            'quarantine_ward_id',
            'care_area',
            'page', 'page_size', 'search', 'order_by',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = UserValidator(**accepted_fields)

            validator.is_valid_fields([
                'status_list', 'health_status_list', 'positive_test_now_list',
                'is_last_tested', 'care_area',
                'created_at_max', 'created_at_min',
            ])
            validator.extra_validate_to_filter_staff()

            query_set = CustomUser.objects.all()

            list_to_filter_staff = [key for key in accepted_fields.keys()]
            list_to_filter_staff = set(list_to_filter_staff) - \
            {'is_last_tested', 'page', 'page_size'}
            list_to_filter_staff = list(list_to_filter_staff) + \
            [
                'status_list',
                'last_tested_max', 'role_name',
            ]

            dict_to_filter_staff = validator.get_data(list_to_filter_staff)

            # Check ward of sender
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if hasattr(validator, '_quarantine_ward'):
                    # Sender want filter with ward, building, floor or room
                    if validator.get_field('quarantine_ward') != request.user.quarantine_ward:
                        raise exceptions.AuthenticationException({'quarantine_ward_id': messages.NO_PERMISSION})
                else:
                    dict_to_filter_staff['quarantine_ward_id'] = request.user.quarantine_ward.id
            
            filter = StaffFilter(dict_to_filter_staff, queryset=query_set)

            query_set = filter.qs

            if 'order_by' not in dict_to_filter_staff.keys():
                query_set = query_set.annotate(num_care_member=Count('member_x_care_staff')).order_by('num_care_member')

            query_set = query_set.select_related('quarantine_ward')

            serializer = FilterStaffSerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class HomeAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='manager', detail=False)
    def manager_home(self, request):
        """Get informations to display home screen for manager

        Args:
            + number_of_days_in_out: int
            - quarantine_ward_id: int
        """

        accept_fields = [
            'number_of_days_in_out',
            'quarantine_ward_id',
        ]

        require_fields = [
            'number_of_days_in_out'
        ]

        try:
            sender_role_name = request.user.role.name
            if sender_role_name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF',]:
                raise exceptions.AuthenticationException()

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            if sender_role_name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if 'quarantine_ward_id' not in accepted_fields.keys():
                    accepted_fields['quarantine_ward_id'] = request.user.quarantine_ward.id
            
            validator = HomeValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['number_of_days_in_out'])
            validator.extra_validate_to_get_manager_home()

            number_of_days_in_out = validator.get_field('number_of_days_in_out')
            quarantine_ward_id = validator.get_field('quarantine_ward_id')
            quarantine_ward = validator.get_field('quarantine_ward')

            if sender_role_name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if int(quarantine_ward_id) != request.user.quarantine_ward.id:
                    raise exceptions.ValidationException({'quarantine_ward_id': messages.NO_PERMISSION})

            users_query_set = CustomUser.objects.all()
            tests_query_set = Test.objects.all()

            # Calculate number of all members past and now

            dict_to_filter_all_members_past_and_now = {
                'role_name': 'MEMBER',
                'status_list': f'{CustomUserStatus.AVAILABLE},{CustomUserStatus.LEAVE}',
            }

            if quarantine_ward_id:
                dict_to_filter_all_members_past_and_now['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_all_members_past_and_now, queryset=users_query_set)

            number_of_all_members_past_and_now = filter.qs.count()

            # Calculate number of quarantining members

            dict_to_filter_quarantining_members = {
                'role_name': 'MEMBER',
                'status': CustomUserStatus.AVAILABLE,
            }

            if quarantine_ward_id:
                dict_to_filter_quarantining_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_quarantining_members, queryset=users_query_set)

            number_of_quarantining_members = filter.qs.count()

            # Calculate number of available slots

            if quarantine_ward_id:
                total_capacity = QuarantineRoom.objects.filter(
                    quarantine_floor__quarantine_building__quarantine_ward__id=quarantine_ward_id
                ).aggregate(Sum('capacity'))['capacity__sum']
            else:
                total_capacity = QuarantineRoom.objects.all().aggregate(Sum('capacity'))['capacity__sum']

            if not total_capacity:
                total_capacity = 0
            number_of_available_slots = total_capacity - number_of_quarantining_members

            # Calculate number of waiting members

            dict_to_filter_waiting_members = {
                'role_name': 'MEMBER',
                'status': CustomUserStatus.WAITING,
            }

            if quarantine_ward_id:
                dict_to_filter_waiting_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_waiting_members, queryset=users_query_set)

            number_of_waiting_members = filter.qs.count()

            # Calculate number of suspected members

            dict_to_filter_suspected_members = {
                'role_name': 'MEMBER',
                'health_status_list': f'{HealthStatus.UNWELL},{HealthStatus.SERIOUS}',
                'positive_test_now_list': [False, None],
                'status': CustomUserStatus.AVAILABLE,
            }

            if quarantine_ward_id:
                dict_to_filter_suspected_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_suspected_members, queryset=users_query_set)

            number_of_suspected_members = filter.qs.count()

            # Calculate number of positive members

            dict_to_filter_positive_members = {
                'role_name': 'MEMBER',
                'positive_test_now_list': [True],
                'status': CustomUserStatus.AVAILABLE,
            }

            if quarantine_ward_id:
                dict_to_filter_positive_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_positive_members, queryset=users_query_set)

            number_of_positive_members = filter.qs.count()

            # Calculate number of need test members

            try:
                pandemic = Pandemic.objects.get(name='Covid-19')
                day_between_tests = int(pandemic.day_between_tests)
            except:
                day_between_tests = int(os.environ.get('DAY_BETWEEN_TESTS', 5))

            last_tested_max = timezone.now() - datetime.timedelta(days=day_between_tests)

            dict_to_filter_need_test_members = {
                'role_name': 'MEMBER',
                'last_tested_max': last_tested_max,
                'status': CustomUserStatus.AVAILABLE,
            }

            if quarantine_ward_id:
                dict_to_filter_need_test_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_need_test_members, queryset=users_query_set)

            number_of_need_test_members = filter.qs.count()

            # Calculate number of can finish members

            dict_to_filter_can_finish_members = {
                'role_name': 'MEMBER',
                'positive_test_now_list': [False],
                'health_status_list': f'{HealthStatus.NORMAL},{HealthStatus.UNWELL}',
                'quarantined_finish_expected_at_max': timezone.now(),
                'status': CustomUserStatus.AVAILABLE,
            }

            if quarantine_ward_id:
                dict_to_filter_can_finish_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_can_finish_members, queryset=users_query_set)

            number_of_can_finish_members = filter.qs.count()

            # Calculate number of hospitalize waiting members

            dict_to_filter_hospitalize_waiting_members = {
                'role_name': 'MEMBER',
                'status': CustomUserStatus.AVAILABLE,
                'quarantined_status': MemberQuarantinedStatus.HOSPITALIZE_WAITING,
            }

            if quarantine_ward_id:
                dict_to_filter_positive_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_hospitalize_waiting_members, queryset=users_query_set)

            number_of_hospitalize_waiting_members = filter.qs.count()

            # Calculate number of completed members

            dict_to_filter_completed_members = {
                'role_name': 'MEMBER',
                'status': CustomUserStatus.LEAVE,
                'quarantined_status': MemberQuarantinedStatus.COMPLETED,
            }

            if quarantine_ward_id:
                dict_to_filter_completed_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_completed_members, queryset=users_query_set)

            number_of_completed_members = filter.qs.count()

            # Calculate number of hospitalized members

            dict_to_filter_hospitalized_members = {
                'role_name': 'MEMBER',
                'status': CustomUserStatus.LEAVE,
                'quarantined_status': MemberQuarantinedStatus.HOSPITALIZE,
            }

            if quarantine_ward_id:
                dict_to_filter_hospitalized_members['quarantine_ward_id'] = quarantine_ward_id

            filter = MemberFilter(dict_to_filter_hospitalized_members, queryset=users_query_set)

            number_of_hospitalized_members = filter.qs.count()

            # Calculate number of waiting tests

            dict_to_filter_waiting_tests = {
                'status': TestStatus.WAITING,
            }

            if quarantine_ward_id:
                dict_to_filter_waiting_tests['quarantine_ward_id'] = quarantine_ward_id

            filter = TestFilter(dict_to_filter_waiting_tests, queryset=tests_query_set)

            number_of_waiting_tests = filter.qs.count()

            # Calculate number of member 'in' everyday

            dict_of_in_members = dict()

            for day_sub in range(number_of_days_in_out - 1, -1, -1):
                day = timezone.now() - datetime.timedelta(days=day_sub)
                day = day.astimezone(pytz.timezone('Asia/Saigon'))
                start_of_day = datetime.datetime(day.year, day.month, day.day)
                start_of_day = start_of_day.astimezone(pytz.timezone('Asia/Saigon'))
                end_of_day = datetime.datetime(day.year, day.month, day.day, 23, 59, 59, 999999)
                end_of_day = end_of_day.astimezone(pytz.timezone('Asia/Saigon'))

                dict_to_filter_in_members = {
                    'role_name': 'MEMBER',
                    'quarantined_at_max': end_of_day,
                    'quarantined_at_min': start_of_day,
                }

                if quarantine_ward_id:
                    dict_to_filter_in_members['quarantine_ward_id'] = quarantine_ward_id

                filter = MemberFilter(dict_to_filter_in_members, queryset=users_query_set)

                dict_of_in_members[f'{day}'[:10]] = filter.qs.count()

            # Calculate number of member 'complete' everyday

            dict_of_complete_members = dict()

            for day_sub in range(number_of_days_in_out - 1, -1, -1):
                day = timezone.now() - datetime.timedelta(days=day_sub)
                day = day.astimezone(pytz.timezone('Asia/Saigon'))
                start_of_day = datetime.datetime(day.year, day.month, day.day)
                start_of_day = start_of_day.astimezone(pytz.timezone('Asia/Saigon'))
                end_of_day = datetime.datetime(day.year, day.month, day.day, 23, 59, 59, 999999)
                end_of_day = end_of_day.astimezone(pytz.timezone('Asia/Saigon'))

                quarantine_history_query_set = QuarantineHistory.objects.filter(
                    status=QuarantineHistoryStatus.ENDED,
                    end_type=QuarantineHistoryEndType.COMPLETED,
                    end_date__lte=end_of_day,
                    end_date__gte=start_of_day,
                )
                
                if quarantine_ward_id:
                    quarantine_history_query_set = quarantine_history_query_set.filter(quarantine_room__quarantine_floor__quarantine_building__quarantine_ward__id=quarantine_ward_id)

                quarantine_history_query_set = quarantine_history_query_set.values('user').distinct()

                dict_of_complete_members[f'{day}'[:10]] = quarantine_history_query_set.count()

            # Calculate number of member 'hospitalize' everyday

            dict_of_hospitalize_members = dict()

            for day_sub in range(number_of_days_in_out - 1, -1, -1):
                day = timezone.now() - datetime.timedelta(days=day_sub)
                day = day.astimezone(pytz.timezone('Asia/Saigon'))
                start_of_day = datetime.datetime(day.year, day.month, day.day)
                start_of_day = start_of_day.astimezone(pytz.timezone('Asia/Saigon'))
                end_of_day = datetime.datetime(day.year, day.month, day.day, 23, 59, 59, 999999)
                end_of_day = end_of_day.astimezone(pytz.timezone('Asia/Saigon'))

                quarantine_history_query_set = QuarantineHistory.objects.filter(
                    status=QuarantineHistoryStatus.ENDED,
                    end_type=QuarantineHistoryEndType.HOSPITALIZE,
                    end_date__lte=end_of_day,
                    end_date__gte=start_of_day,
                )
                
                if quarantine_ward_id:
                    quarantine_history_query_set = quarantine_history_query_set.filter(quarantine_room__quarantine_floor__quarantine_building__quarantine_ward__id=quarantine_ward_id)

                quarantine_history_query_set = quarantine_history_query_set.values('user').distinct()

                dict_of_hospitalize_members[f'{day}'[:10]] = quarantine_history_query_set.count()

            if quarantine_ward:
                serializer = BaseQuarantineWardSerializer(quarantine_ward, many=False)
                quarantine_ward_data = serializer.data
            else:
                quarantine_ward_data = 'All'

            response_data = {
                'quarantine_ward': quarantine_ward_data,
                'number_of_all_members_past_and_now': number_of_all_members_past_and_now,
                'number_of_quarantining_members': number_of_quarantining_members,
                'number_of_available_slots': number_of_available_slots,
                'number_of_waiting_members': number_of_waiting_members,
                'number_of_suspected_members': number_of_suspected_members,
                'number_of_positive_members': number_of_positive_members,
                'number_of_need_test_members': number_of_need_test_members,
                'number_of_can_finish_members': number_of_can_finish_members,
                'number_of_hospitalize_waiting_members': number_of_hospitalize_waiting_members,
                'number_of_completed_members': number_of_completed_members,
                'number_of_hospitalized_members': number_of_hospitalized_members,
                'number_of_waiting_tests': number_of_waiting_tests,
                'in': dict_of_in_members,
                'complete': dict_of_complete_members,
                'hospitalize': dict_of_hospitalize_members,
            }

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)
    
    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter_address_with_num_of_members_pass_by', detail=False)
    def filter_address_with_num_of_members_pass_by(self, request):
        """Get a list of cities/districts/wards and number of members pass by it

        Args:
            + address_type: String ['city', 'district', 'ward']
            - father_address_id: int
            - quarantine_ward_id: int
            - start_time_max: String vd:'2000-01-26T01:23:45.123456Z'
            - start_time_min: String vd:'2000-01-26T01:23:45.123456Z'
            - page: int
            - page_size: int
            - search: String
            - order_by: String ['name', 'num_of_members_pass_by'], mặc định sắp thứ tự theo num_of_members_pass_by giảm dần
        """

        accept_fields = [
            'address_type', 'father_address_id',
            'quarantine_ward_id', 'start_time_max', 'start_time_min',
            'page', 'page_size', 'search', 'order_by',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})
                
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = HomeValidator(**accepted_fields)
            validator.is_missing_fields(['address_type'])
            validator.is_valid_fields(['address_type', 'start_time_max', 'start_time_min'])
            validator.extra_validate_to_filter_address_with_num_of_members_pass_by()

            address_type = validator.get_field('address_type')
            father_address_id = validator.get_field('father_address_id')
            quarantine_ward = validator.get_field('quarantine_ward')
            start_time_max = validator.get_field('start_time_max')
            start_time_min = validator.get_field('start_time_min')
            search_value = validator.get_field('search')
            order_by = validator.get_field('order_by')

            response_list = []

            if address_type == 'city':
                all_addresses = City.objects.filter(country__code='VNM')
            elif address_type == 'district':
                all_addresses = District.objects.filter(city__id=father_address_id)
            else:
                all_addresses = Ward.objects.filter(district__id=father_address_id)
            if search_value:
                all_addresses = all_addresses.filter(name__unaccent__icontains=search_value)

            for address in list(all_addresses):
                query_set = DestinationHistory.objects.filter(
                    user__status__in=[CustomUserStatus.AVAILABLE, CustomUserStatus.LEAVE],
                    user__role__name='MEMBER',
                )
                if address_type == 'city':
                    query_set = query_set.filter(city=address)
                elif address_type == 'district':
                    query_set = query_set.filter(district=address)
                else:
                    query_set = query_set.filter(ward=address)
                if quarantine_ward:
                    query_set = query_set.filter(user__quarantine_ward=quarantine_ward)
                if start_time_max:
                    query_set = query_set.filter(start_time__lte=start_time_max)
                if start_time_min:
                    query_set = query_set.filter(start_time__gte=start_time_min)
                num_of_members_pass_by = query_set.values('user').annotate(num_user=Count('user')).count()
                response_list += [{
                    address_type: {
                        'id': address.id,
                        'name': address.name,
                    },
                    'num_of_members_pass_by': num_of_members_pass_by,
                }]

            if order_by:
                locale.setlocale(locale.LC_COLLATE, '')
                if order_by == 'name':
                    key_to_sort = lambda d: locale.strxfrm(d[address_type]['name'])
                    reverse_to_sort = False
                elif order_by == '-name':
                    key_to_sort = lambda d: locale.strxfrm(d[address_type]['name'])
                    reverse_to_sort = True
                elif order_by == 'num_of_members_pass_by':
                    key_to_sort = lambda d: d['num_of_members_pass_by']
                    reverse_to_sort = False
                elif order_by == '-num_of_members_pass_by':
                    key_to_sort = lambda d: d['num_of_members_pass_by']
                    reverse_to_sort = True
            else:
                key_to_sort = lambda d: d['num_of_members_pass_by']
                reverse_to_sort = True

            response_list = sorted(response_list, key=key_to_sort, reverse=reverse_to_sort)

            paginated_data = paginate_data(request, response_list)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='member', detail=False)
    def member_home(self, request):
        """Get information to display home screen for member

        Args:
            None
        """

        try:
            if request.user.role.name not in ['MEMBER']:
                raise exceptions.AuthenticationException()

            if hasattr(request.user, 'member_x_custom_user'):
                member = request.user.member_x_custom_user
                member = Member.objects.select_related(
                    'custom_user__quarantine_ward__main_manager',
                    'custom_user__quarantine_ward__country',
                    'custom_user__quarantine_ward__city',
                    'custom_user__quarantine_ward__district',
                    'custom_user__quarantine_ward__ward',
                    'custom_user__quarantine_ward__pandemic',
                    'quarantine_room__quarantine_floor__quarantine_building',
                    'care_staff',
                ).get(id=member.id)
            else:
                member = None

            serializer = MemberHomeSerializer(member, many=False)

            return_data = serializer.data

            if member:
                if member.custom_user.status == CustomUserStatus.LEAVE and member.quarantined_status == MemberQuarantinedStatus.HOSPITALIZE:
                    # if this member is hospitalize
                    hospital_name = None
                    note = None
                    time = None
                    quarantine_histories = list(QuarantineHistory.objects.filter(user=member.custom_user).order_by('-start_date'))
                    if len(quarantine_histories) >= 1:
                        last_quarantine_history = quarantine_histories[0]
                        if last_quarantine_history.status == QuarantineHistoryStatus.ENDED and last_quarantine_history.end_type == QuarantineHistoryEndType.HOSPITALIZE:
                            temp = last_quarantine_history.note.split(";") if last_quarantine_history.note else []
                            hospital_name = temp[0] if len(temp) >= 1 else None
                            note = temp[1] if len(temp) >= 2 else None
                            time = last_quarantine_history.end_date
                    return_data['hospitalize_info'] = {
                        'hospital_name': hospital_name,
                        'note': note,
                        'time': time,
                    }

            return self.response_handler.handle(data=return_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)
