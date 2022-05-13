import os
import datetime
import pytz
import openpyxl, csv, codecs
from random import randint
from django.db.models import Q
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from utils.views import paginate_data
from rest_framework import permissions
from rest_framework.parsers import MultiPartParser
from rest_framework.decorators import action
from drf_yasg.utils import swagger_auto_schema
from .models import BackgroundDisease, MedicalDeclaration, Symptom, Test, Vaccine, VaccineDose, Pandemic
from .validators.medical_declaration import MedicalDeclarationValidator
from .validators.test import TestValidator
from .validators.vaccine import VaccineValidator, VaccineDoseValidator
from .validators.pandemic import PandemicValidator
from .serializers import (
    PandemicSerializer,
    MedicalDeclarationSerializer,
    FilterMedicalDeclarationSerializer,
    TestSerializer,
    FilterTestSerializer,
    BaseBackgroundDiseaseSerializer,
    BaseSymptomSerializer,
    VaccineSerializer,
    VaccineDoseSerializer,
    FilterVaccineDoseSerializer,
)
from .filters.medical_declaration import MedicalDeclarationFilter
from .filters.test import TestFilter
from .filters.vaccine import VaccineDoseFilter
from .swagger_params import (
    get_pandemic_swagger_params,
    update_pandemic_swagger_params,
)
from user_account.serializers import (
    BaseBaseCustomUserSerializer,
)
from notification.views import create_and_send_noti_to_list_user
from utils import exceptions, messages
from utils.enums import SymptomType, TestResult, TestType, HealthStatus, MemberLabel, CustomUserStatus
from utils.views import AbstractView, query_debugger

# Create your views here.

class PandemicAPI(AbstractView):

    parser_classes = (MultiPartParser,)
    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_pandemic(self, request):
        """Create a pandemic

        Args:
            + name: String
        """

        accept_fields = [
            'name',
        ]

        require_fields = [
            'name',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = PandemicValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_create_pandemic()

            list_to_create_pandemic = [key for key in accepted_fields.keys()]

            dict_to_create_pandemic = validator.get_data(list_to_create_pandemic)

            pandemic = Pandemic(**dict_to_create_pandemic)
            
            pandemic.created_by = request.user
            pandemic.updated_by = request.user
            
            pandemic.save()

            serializer = PandemicSerializer(pandemic, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @swagger_auto_schema(method='post', manual_parameters=get_pandemic_swagger_params, responses={200: ''})
    @action(methods=['POST'], url_path='get', detail=False)
    def get_pandemic(self, request):
        """Get a pandemic

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = PandemicValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_to_get_pandemic()

            pandemic = validator.get_field('pandemic')

            serializer = PandemicSerializer(pandemic, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @swagger_auto_schema(method='post', manual_parameters=update_pandemic_swagger_params, responses={200: ''})
    @action(methods=['POST'], url_path='update', detail=False)
    def update_pandemic(self, request):
        """Update a pandemic

        Args:
            + id: int
            - name: String
            - quarantine_time_not_vac: int
            - quarantine_time_vac: int
            - remain_qt_cc_pos_vac: int
            - remain_qt_cc_pos_not_vac: int
            - remain_qt_cc_not_pos_vac: int
            - remain_qt_cc_not_pos_not_vac: int
            - remain_qt_pos_vac: int
            - remain_qt_pos_not_vac: int
            - test_type_pos_to_neg_vac: String ['QUICK', 'RT-PCR']
            - num_test_pos_to_neg_vac: int
            - test_type_pos_to_neg_not_vac: String ['QUICK', 'RT-PCR']
            - num_test_pos_to_neg_not_vac: int
            - test_type_none_to_neg_vac: String ['QUICK', 'RT-PCR']
            - num_test_none_to_neg_vac: int
            - test_type_none_to_neg_not_vac: String ['QUICK', 'RT-PCR']
            - num_test_none_to_neg_not_vac: int
            - num_day_to_close_room: int
            - day_between_tests: int
        """

        accept_fields = [
            'id', 'name', 'quarantine_time_not_vac',
            'quarantine_time_vac', 'remain_qt_cc_pos_vac',
            'remain_qt_cc_pos_not_vac', 'remain_qt_cc_not_pos_vac',
            'remain_qt_cc_not_pos_not_vac', 'remain_qt_pos_vac',
            'remain_qt_pos_not_vac', 'test_type_pos_to_neg_vac',
            'num_test_pos_to_neg_vac', 'test_type_pos_to_neg_not_vac',
            'num_test_pos_to_neg_not_vac', 'test_type_none_to_neg_vac',
            'num_test_none_to_neg_vac', 'test_type_none_to_neg_not_vac',
            'num_test_none_to_neg_not_vac', 'num_day_to_close_room',
            'day_between_tests',
        ]

        require_fields = [
            'id',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = PandemicValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.extra_validate_positive_integer([
                'quarantine_time_not_vac', 'quarantine_time_vac',
                'remain_qt_cc_pos_vac', 'remain_qt_cc_pos_not_vac',
                'remain_qt_cc_not_pos_vac', 'remain_qt_cc_not_pos_not_vac',
                'remain_qt_pos_vac', 'remain_qt_pos_not_vac',
                'num_test_pos_to_neg_vac', 'num_test_pos_to_neg_not_vac',
                'num_test_none_to_neg_vac', 'num_test_none_to_neg_not_vac',
                'num_day_to_close_room', 'day_between_tests',
            ])
            validator.extra_validate_test_type([
                'test_type_pos_to_neg_vac', 'test_type_pos_to_neg_not_vac',
                'test_type_none_to_neg_vac', 'test_type_none_to_neg_not_vac',
            ])
            validator.extra_validate_to_update_pandemic()

            pandemic = validator.get_field('pandemic')

            list_to_update_pandemic = [key for key in accepted_fields.keys()]
            list_to_update_pandemic = set(list_to_update_pandemic) - {'id'}

            dict_to_update_pandemic = validator.get_data(list_to_update_pandemic)

            for attr, value in dict_to_update_pandemic.items(): 
                setattr(pandemic, attr, value)

            pandemic.updated_by = request.user
            pandemic.save()

            serializer = PandemicSerializer(pandemic, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_pandemic(self, request):
        """Get a list of pandemics

        Args:
            None    
        """

        try:

            pandemics = Pandemic.objects.all()

            serializer = PandemicSerializer(pandemics, many=True)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class BackgroundDiseaseAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='init', detail=False)
    def init_background_disease(self, request):
        """Init some background disease

        Args:
            None
        """

        try:
            user = request.user
            if not user.role.name == 'ADMINISTRATOR':
                raise exceptions.AuthenticationException()

            disease_list = [
                'Tiểu đường', 'Ung thư', 'Hen suyễn', 'Tăng huyết áp',
                'Bệnh gan', 'Bệnh thận mãn tính', 'Bệnh tim mạch',
                'Bệnh lý mạch máu não', 'Bệnh khác',
            ]
            
            for name in disease_list:
                try:
                    disease = BackgroundDisease.objects.get(name=name)
                except:
                    disease = BackgroundDisease(name=name)
                    disease.save()

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_background_disease(self, request):
        """List all background disease

        Args:
            None
        """

        try:
            serializer = BaseBackgroundDiseaseSerializer(BackgroundDisease.objects.all(), many=True)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)


class SymptomAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='init', detail=False)
    def init_symptom(self, request):
        """Init some symptom
        Args:
            None
        """

        try:
            user = request.user
            if not user.role.name == 'ADMINISTRATOR':
                raise exceptions.AuthenticationException()

            main_symptom_list = [
                'Ho ra máu', 'Thở dốc, khó thở', 'Đau tức ngực kéo dài', 'Lơ mơ, không tỉnh táo',
            ]

            extra_symptom_list = [
                'Mệt mỏi', 'Ho', 'Ho có đờm', 'Đau họng', 'Đau đầu', 'Chóng mặt',
                'Chán ăn', 'Nôn / Buồn nôn', 'Tiêu chảy', 'Xuất huyết ngoài da',
                'Nổi ban ngoài da', 'Ớn lạnh / gai rét', 'Viêm kết mạc (mắt đỏ)',
                'Mất vị giác, khứu giác', 'Đau nhức cơ',
            ]
            
            for name in main_symptom_list:
                try:
                    disease = Symptom.objects.get(name=name, type=SymptomType.MAIN)
                except:
                    disease = Symptom(name=name, type=SymptomType.MAIN)
                    disease.save()

            for name in extra_symptom_list:
                try:
                    disease = Symptom.objects.get(name=name, type=SymptomType.EXTRA)
                except:
                    disease = Symptom(name=name, type=SymptomType.EXTRA)
                    disease.save()

            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_symptom(self, request):
        """List all symptom

        Args:
            None
        """

        try:
            response_data = dict()

            serializer = BaseSymptomSerializer(Symptom.objects.filter(type=SymptomType.MAIN), many=True)
            response_data['main'] = serializer.data

            serializer = BaseSymptomSerializer(Symptom.objects.filter(type=SymptomType.EXTRA), many=True)
            response_data['extra'] = serializer.data

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class MedicalDeclarationAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    def custom_medical_declaration_code_generator(self, user_code):
        first_part_length = int(os.environ.get("MEDICAL_DECLARATION_CODE_USER_CODE_LENGTH", "3"))
        first_part = ('0000000000' + str(user_code))[-first_part_length:]
        
        second_part_length = int(os.environ.get("MEDICAL_DECLARATION_CODE_TIMESTAMP_LENGTH", "6"))
        second_part = ('0000000000' + str(int(datetime.datetime.now().timestamp())))[-second_part_length:]

        third_part_length = int(os.environ.get("MEDICAL_DECLARATION_CODE_RANDOM_LENGTH", "6"))
        third_part = ''.join(str(randint(0, 9)) for i in range(third_part_length))

        return first_part + second_part + third_part

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_medical_declaration(self, request):
        """Create a medical declaration

        Args:
            - phone_number: String
            - heartbeat: int
            - temperature: float
            - breathing: int
            - spo2: float
            - blood_pressure_max: int
            - blood_pressure_min: int
            - main_symptoms: String '<id>,<id>,<id>'
            - extra_symptoms: String '<id>,<id>,<id>'
            - other_symptoms: String
        """

        accept_fields = [
            'phone_number', 'heartbeat', 'temperature', 
            'breathing', 'spo2',
            'blood_pressure_max',
            'blood_pressure_min',
            'main_symptoms', 'extra_symptoms', 'other_symptoms',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = MedicalDeclarationValidator(**accepted_fields)
            validator.is_valid_fields([
                'phone_number', 'heartbeat', 'temperature',
                'breathing', 'spo2',
                'blood_pressure_max', 'blood_pressure_min',
                'main_symptoms', 'extra_symptoms',
            ])
            
            validator.extra_validate_to_create_medical_declaration()

            list_to_create_medical_declaration = [key for key in accepted_fields.keys()]
            list_to_create_medical_declaration = set(list_to_create_medical_declaration) - {'phone_number'}
            list_to_create_medical_declaration = list(list_to_create_medical_declaration) + ['conclude']

            dict_to_create_medical_declaration = validator.get_data(list_to_create_medical_declaration)

            medical_declaration = MedicalDeclaration(**dict_to_create_medical_declaration)
            
            for_user = request.user
            if 'phone_number' in accepted_fields.keys():
                for_user = validator.get_field('user')

            medical_declaration.user = for_user

            medical_declaration.code = self.custom_medical_declaration_code_generator(medical_declaration.user.code)
            while (validator.is_code_exist(medical_declaration.code)):
                medical_declaration.code = self.custom_medical_declaration_code_generator(medical_declaration.user.code)

            medical_declaration.created_by = request.user
            medical_declaration.save()

            if hasattr(for_user, 'member_x_custom_user'):
                for_member = for_user.member_x_custom_user
                for_member.health_status = medical_declaration.conclude
                for_member.save()

            if hasattr(for_user, 'manager_x_custom_user'):
                for_manager = for_user.manager_x_custom_user
                for_manager.health_status = medical_declaration.conclude
                for_manager.save()
            
            if hasattr(for_user, 'staff_x_custom_user'):
                for_staff = for_user.staff_x_custom_user
                for_staff.health_status = medical_declaration.conclude
                for_staff.save()

            if medical_declaration.conclude == HealthStatus.SERIOUS and medical_declaration.user.role.name == 'MEMBER' \
            and hasattr(medical_declaration.user, 'member_x_custom_user') and medical_declaration.user.status == CustomUserStatus.AVAILABLE:
                # Send notification to care_staff or manager
                receive_user = medical_declaration.user.member_x_custom_user.care_staff
                if not receive_user:
                    receive_user = medical_declaration.user.quarantine_ward.main_manager
                if receive_user:
                    title = 'Cảnh báo sức khỏe người cách ly'
                    description = f'Người cách ly {medical_declaration.user.full_name} ở {medical_declaration.user.member_x_custom_user.quarantine_room.name} - ' + \
                                f'{medical_declaration.user.member_x_custom_user.quarantine_floor.name} - ' + \
                                f'{medical_declaration.user.member_x_custom_user.quarantine_building.name} - ' + \
                                f'{medical_declaration.user.member_x_custom_user.quarantine_ward.full_name} ' + \
                                f'đang có dấu hiệu sức khỏe nguy hiểm'
                    create_and_send_noti_to_list_user(title, description, created_by=medical_declaration.user, receive_user_list=[receive_user])

            serializer = MedicalDeclarationSerializer(medical_declaration, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_medical_declaration(self, request):
        """Get a medical declaration

        Args:
            + id: String
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = MedicalDeclarationValidator(**accepted_fields)

            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_get_medical_declaration()

            medical_declaration = validator.get_field('medical_declaration')

            serializer = MedicalDeclarationSerializer(medical_declaration, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get_health_info_of_user', detail=False)
    def get_health_info_of_user(self, request):
        """Get all health info of a user

        Args:
            - user_code: String
        """

        accept_fields = [
            'user_code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            if 'user_code' not in accepted_fields.keys() or not accepted_fields['user_code']:
                accepted_fields['user_code'] = request.user.code
            validator = MedicalDeclarationValidator(**accepted_fields)
            validator.extra_validate_to_get_health_info_of_user()

            user = validator.get_field('user')

            response_dict = {
                'user': BaseBaseCustomUserSerializer(user, many=False).data,
                'heartbeat': None,
                'temperature': None,
                'breathing': None,
                'spo2': None,
                'blood_pressure_max': None,
                'blood_pressure_min': None,
                'main_symptoms': None,
                'extra_symptoms': None,
                'other_symptoms': None,
            }

            last_medical_declaration = MedicalDeclaration.objects.filter(user=user).order_by('created_at').last()
            if last_medical_declaration:
                main_symptoms = {
                    'data': last_medical_declaration.main_symptoms,
                    'updated_at': last_medical_declaration.created_at,
                }
                extra_symptoms = {
                    'data': last_medical_declaration.extra_symptoms,
                    'updated_at': last_medical_declaration.created_at,
                }
                other_symptoms = {
                    'data': last_medical_declaration.other_symptoms,
                    'updated_at': last_medical_declaration.created_at,
                }
                response_dict['main_symptoms'] = main_symptoms
                response_dict['extra_symptoms'] = extra_symptoms
                response_dict['other_symptoms'] = other_symptoms
            
            heartbeat_medical_declaration = MedicalDeclaration.objects.filter(user=user, heartbeat__isnull=False).order_by('created_at').last()
            if heartbeat_medical_declaration:
                heartbeat = {
                    'data': heartbeat_medical_declaration.heartbeat,
                    'updated_at': heartbeat_medical_declaration.created_at,
                }
                response_dict['heartbeat'] = heartbeat

            temperature_medical_declaration = MedicalDeclaration.objects.filter(user=user, temperature__isnull=False).order_by('created_at').last()
            if temperature_medical_declaration:
                temperature = {
                    'data': temperature_medical_declaration.temperature,
                    'updated_at': temperature_medical_declaration.created_at,
                }
                response_dict['temperature'] = temperature

            breathing_medical_declaration = MedicalDeclaration.objects.filter(user=user, breathing__isnull=False).order_by('created_at').last()
            if breathing_medical_declaration:
                breathing = {
                    'data': breathing_medical_declaration.breathing,
                    'updated_at': breathing_medical_declaration.created_at,
                }
                response_dict['breathing'] = breathing

            spo2_medical_declaration = MedicalDeclaration.objects.filter(user=user, spo2__isnull=False).order_by('created_at').last()
            if spo2_medical_declaration:
                spo2 = {
                    'data': spo2_medical_declaration.spo2,
                    'updated_at': spo2_medical_declaration.created_at,
                }
                response_dict['spo2'] = spo2

            blood_pressure_max_medical_declaration = MedicalDeclaration.objects.filter(user=user, blood_pressure_max__isnull=False).order_by('created_at').last()
            if blood_pressure_max_medical_declaration:
                blood_pressure_max = {
                    'data': blood_pressure_max_medical_declaration.blood_pressure_max,
                    'updated_at': blood_pressure_max_medical_declaration.created_at,
                }
                response_dict['blood_pressure_max'] = blood_pressure_max

            blood_pressure_min_medical_declaration = MedicalDeclaration.objects.filter(user=user, blood_pressure_min__isnull=False).order_by('created_at').last()
            if blood_pressure_min_medical_declaration:
                blood_pressure_min = {
                    'data': blood_pressure_min_medical_declaration.blood_pressure_min,
                    'updated_at': blood_pressure_min_medical_declaration.created_at,
                }
                response_dict['blood_pressure_min'] = blood_pressure_min

            return self.response_handler.handle(data=response_dict)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_medical_declaration(self, request):
        """Get a list of medical declarations

        Args:
            - user_code: String
            - created_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - created_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - page: int
            - page_size: int
            - search: String
        """

        accept_fields = [
            'user_code', 'page', 'page_size', 'search',
            'created_at_max', 'created_at_min',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = MedicalDeclarationValidator(**accepted_fields)

            validator.is_valid_fields([
                'created_at_max', 'created_at_min',
            ])
            validator.extra_validate_to_filter_medical_declaration()

            query_set = MedicalDeclaration.objects.all()

            list_to_filter_medical_declaration = [key for key in accepted_fields.keys()]
            list_to_filter_medical_declaration = set(list_to_filter_medical_declaration) - \
            {'page', 'page_size'}

            dict_to_filter_medical_declaration = validator.get_data(list_to_filter_medical_declaration)

            dict_to_filter_medical_declaration.setdefault('order_by', '-created_at')

            filter = MedicalDeclarationFilter(dict_to_filter_medical_declaration, queryset=query_set)

            query_set = filter.qs

            query_set = query_set.select_related('user__member_x_custom_user', 'user__manager_x_custom_user', 'user__staff_x_custom_user')

            serializer = FilterMedicalDeclarationSerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class TestAPI(AbstractView):
    permission_classes = [permissions.IsAuthenticated]

    def custom_test_code_generator(self, user_code):
        first_part_length = int(os.environ.get("TEST_CODE_USER_CODE_LENGTH", "3"))
        first_part = ('0000000000' + str(user_code))[-first_part_length:]
        
        second_part_length = int(os.environ.get("TEST_CODE_TIMESTAMP_LENGTH", "6"))
        second_part = ('0000000000' + str(int(datetime.datetime.now().timestamp())))[-second_part_length:]

        third_part_length = int(os.environ.get("TEST_CODE_RANDOM_LENGTH", "6"))
        third_part = ''.join(str(randint(0, 9)) for i in range(third_part_length))

        return first_part + second_part + third_part

    def calculate_new_conclude_from_test(self, test, old_positive_test_now):
        """test must be the last test has result of this member"""
        if test.result == TestResult.POSITIVE:
            return True
        elif test.result == TestResult.NEGATIVE:
            if old_positive_test_now == False or old_positive_test_now == None:
                return False
            else:
                this_user = test.user
                quarantine_ward = this_user.quarantine_ward
                number_of_vaccine_doses = 0
                if hasattr(this_user, 'member_x_custom_user'):
                    number_of_vaccine_doses = this_user.member_x_custom_user.number_of_vaccine_doses
                elif hasattr(this_user, 'staff_x_custom_user'):
                    number_of_vaccine_doses = this_user.staff_x_custom_user.number_of_vaccine_doses
                elif hasattr(this_user, 'manager_x_custom_user'):
                    number_of_vaccine_doses = this_user.manager_x_custom_user.number_of_vaccine_doses

                # get number_test_need and type_test_need
                if quarantine_ward.pandemic:
                    if number_of_vaccine_doses < 2:
                        number_test_need = quarantine_ward.pandemic.num_test_pos_to_neg_not_vac
                        type_test_need = quarantine_ward.pandemic.test_type_pos_to_neg_not_vac
                    else:
                        number_test_need = quarantine_ward.pandemic.num_test_pos_to_neg_vac
                        type_test_need = quarantine_ward.pandemic.test_type_pos_to_neg_vac
                else:
                    if number_of_vaccine_doses < 2:
                        number_test_need = int(os.environ.get('NUM_TEST_POS_TO_NEG_NOT_VAC', 1))
                        type_test_need = os.environ.get('TEST_TYPE_POS_TO_NEG_NOT_VAC', TestType.RT_PCR)
                    else:
                        number_test_need = int(os.environ.get('NUM_TEST_POS_TO_NEG_VAC', 1))
                        type_test_need = os.environ.get('TEST_TYPE_POS_TO_NEG_VAC', TestType.QUICK)

                # check there tests
                tests_to_check = Test.objects.filter(user=this_user)
                if type_test_need == TestType.RT_PCR:
                    tests_to_check = tests_to_check.filter(type=type_test_need)
                tests_to_check = tests_to_check.filter(~Q(result=TestResult.NONE))
                positive_tests_of_this_user = Test.objects.filter(user=this_user, result=TestResult.POSITIVE)
                if positive_tests_of_this_user.count() >= 1:
                    last_positive_test_of_this_user = positive_tests_of_this_user.order_by('-created_at')[:1].get()
                    tests_to_check = tests_to_check.filter(created_at__gt=last_positive_test_of_this_user.created_at)
                
                tests_to_check = tests_to_check.order_by('-created_at')[:number_test_need]
                tests_to_check = list(tests_to_check)
                if len(tests_to_check) < number_test_need:
                    return True
                is_positive = False
                for test in tests_to_check:
                    if test.result == TestResult.POSITIVE:
                        is_positive = True
                        break
                return is_positive
        else:
            return old_positive_test_now

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_test(self, request):
        """Create a test

        Args:
            + phone_number: String
            + type: String ['QUICK', 'RT-PCR']
            + result: String ['NONE', 'NEGATIVE', 'POSITIVE']
        """

        accept_fields = [
            'phone_number', 'type',
            'result',
        ]

        require_fields = [
            'phone_number', 'type',
            'result',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = TestValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'type', 'result',
            ])
            
            validator.extra_validate_to_create_test()

            list_to_create_test = [key for key in accepted_fields.keys()]
            list_to_create_test = set(list_to_create_test) - {'phone_number'}
            list_to_create_test = list(list_to_create_test) + ['user', 'status']

            dict_to_create_test = validator.get_data(list_to_create_test)

            test = Test(**dict_to_create_test)

            test.code = self.custom_test_code_generator(test.user.code)
            while (validator.is_code_exist(test.code)):
                test.code = self.custom_test_code_generator(test.user.code)
            test.created_by = request.user
            test.save()

            # Update user
            user = test.user
            if hasattr(user, 'member_x_custom_user'):
                this_member = user.member_x_custom_user
                this_member.last_tested = test.created_at

                if test.result != TestResult.NONE:
                    old_positive_test_now = this_member.positive_test_now
                    new_positive_test_now = self.calculate_new_conclude_from_test(test, old_positive_test_now)
                    this_member.positive_test_now = new_positive_test_now

                    if old_positive_test_now != True and new_positive_test_now == True:
                        this_member.label = MemberLabel.F0
                        if not this_member.first_positive_test_date:
                            this_member.first_positive_test_date = test.created_at

                        if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                            if this_member.quarantined_finish_expected_at == None:
                                quarantine_ward = this_member.custom_user.quarantine_ward
                                if quarantine_ward.pandemic:
                                    if this_member.number_of_vaccine_doses < 2:
                                        remain_qt = quarantine_ward.pandemic.remain_qt_pos_not_vac
                                    else:
                                        remain_qt = quarantine_ward.pandemic.remain_qt_pos_vac
                                else:
                                    if this_member.number_of_vaccine_doses < 2:
                                        remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                                    else:
                                        remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                                this_member.quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                            
                            # affect other member in this room
                            this_room = this_member.quarantine_room
                            if this_room:
                                members_in_this_room = this_room.member_x_quarantine_room.all()
                                for member in list(members_in_this_room):
                                    if member.label != MemberLabel.F0:
                                        member.label = MemberLabel.F1
                                        member.quarantined_finish_expected_at = None
                                        member.save()

                    elif old_positive_test_now == True and new_positive_test_now == False:
                        this_member.positive_tested_before = True

                        if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                            # affect other member in this room
                            this_room = this_member.quarantine_room
                            if this_room:
                                other_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=this_member.id)
                                number_of_other_positive_member_in_this_room = other_members_in_this_room.filter(positive_test_now=True).count()
                                if number_of_other_positive_member_in_this_room == 0:
                                    quarantine_ward = this_room.quarantine_floor.quarantine_building.quarantine_ward
                                    for member in list(other_members_in_this_room):
                                        if member.label != MemberLabel.F0:
                                            if quarantine_ward.pandemic:
                                                if member.number_of_vaccine_doses < 2:
                                                    remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_not_vac
                                                else:
                                                    remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_vac
                                            else:
                                                if member.number_of_vaccine_doses < 2:
                                                    remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_NOT_VAC', 14))
                                                else:
                                                    remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_VAC', 10))
                                            old_quarantined_finish_expected_at = member.quarantined_finish_expected_at
                                            new_quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                                            if not old_quarantined_finish_expected_at or old_quarantined_finish_expected_at < new_quarantined_finish_expected_at:
                                                member.quarantined_finish_expected_at = new_quarantined_finish_expected_at
                                                member.save()

                    this_member.last_tested_had_result = test.created_at
                this_member.save()
            if hasattr(user, 'manager_x_custom_user'):
                this_manager = user.manager_x_custom_user
                this_manager.last_tested = test.created_at
                if test.result != TestResult.NONE:
                    new_positive_test_now = self.calculate_new_conclude_from_test(test, this_manager.positive_test_now)
                    this_manager.positive_test_now = new_positive_test_now
                    this_manager.last_tested_had_result = test.created_at
                this_manager.save()
            if hasattr(user, 'staff_x_custom_user'):
                this_staff = user.staff_x_custom_user
                this_staff.last_tested = test.created_at
                if test.result != TestResult.NONE:
                    new_positive_test_now = self.calculate_new_conclude_from_test(test, this_staff.positive_test_now)
                    this_staff.positive_test_now = new_positive_test_now
                    this_staff.last_tested_had_result = test.created_at
                this_staff.save()

            # Send notification to this user
            if test.result != TestResult.NONE:
                vntz = pytz.timezone('Asia/Saigon')
                created_at = test.created_at.astimezone(vntz)
                time_string = f'{created_at.hour} giờ {created_at.minute} phút {created_at.second} giây, ngày {created_at.day} tháng {created_at.month} năm {created_at.year}'
                result_string = 'ÂM TÍNH' if test.result == TestResult.NEGATIVE else 'DƯƠNG TÍNH'
                title = 'Kết quả xét nghiệm'
                description = 'Phiếu xét nghiệm lúc ' + time_string + ' có kết quả ' + result_string
                create_and_send_noti_to_list_user(title=title, description=description, receive_user_list=[test.user], created_by=None)

            serializer = TestSerializer(test, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)
        
    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create_many', detail=False)
    def create_many_test(self, request):
        """Create many test

        Args:
            + user_codes: String <code>,<code>
            + type: String ['QUICK', 'RT-PCR']
        """

        accept_fields = [
            'user_codes', 'type',
        ]

        require_fields = [
            'user_codes', 'type',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = TestValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'type',
            ])
            
            validator.extra_validate_to_create_many_test()

            users = validator.get_field('users')
            type = validator.get_field('type')

            tests = []
            for user in users:
                test = Test(user=user, type=type, created_by=request.user, updated_by=request.user)
                test.code = self.custom_test_code_generator(user.code)
                while (validator.is_code_exist(test.code)):
                    test.code = self.custom_test_code_generator(user.code)
                tests += [test]

            Test.objects.bulk_create(tests)

            # Update user
            for test in tests:
                user = test.user
                if hasattr(user, 'member_x_custom_user'):
                    this_member = user.member_x_custom_user
                    this_member.last_tested = test.created_at
                    this_member.save()
                if hasattr(user, 'manager_x_custom_user'):
                    this_manager = user.manager_x_custom_user
                    this_manager.last_tested = test.created_at
                    this_manager.save()
                if hasattr(user, 'staff_x_custom_user'):
                    this_staff = user.staff_x_custom_user
                    this_staff.last_tested = test.created_at
                    this_staff.save()

            return self.response_handler.handle(message=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create_by_file', detail=False)
    def create_test_by_file(self, request):
        """Create a test by file (csv, xlsx)
        
        Args:
            + file: .csv, .xlsx
        """

        test_fields = [
            'user_code', 'status', 'type',
            'result',
        ]

        status_switcher = {
            "Chờ kết quả": 'WAITING',
            "Đã có kết quả": 'DONE',
        }

        type_switcher = {
            "Test nhanh": 'QUICK',
            "Test Real-time PCR": 'RT-PCR'
        }

        result_switcher = {
            "Chưa có": 'NONE',
            "Âm tính": 'NEGATIVE',
            "Dương tính": 'POSITIVE',
        }

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            file_name = request.FILES.get('file', False)
            if file_name:
                extension = os.path.splitext(str(file_name))[1]
                iter_rows = None
                error = dict()
                test_ids = list()
                is_csv_file = False
                list_user_positive = list()
                list_user_negative = list()

                if (extension == ".xlsx"):
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb["Sheet1"]
                    iter_rows = ws.iter_rows(min_row=2, min_col=2, max_col=5)
                elif (extension == ".csv"):
                    iter_rows = csv.DictReader(codecs.iterdecode(file_name, encoding='utf-8'))
                    is_csv_file = True

                for row_index, row in enumerate(iter_rows):
                    try:
                        dict_test_data = dict()
                        check_data = None

                        if is_csv_file: 
                            row = list(row.values())[1:]
                            check_data = row[0]
                        else:
                            check_data = row[0].value
                        if check_data in [None, '']:
                            continue

                        for index, data in enumerate(row):
                            value = data.value if not is_csv_file else data
                            if index == 1:
                                value = str(value)
                                value = status_switcher[value]
                            elif index == 2:
                                value = str(value)
                                value = type_switcher[value]
                            elif index == 3:
                                value = str(value)
                                value = result_switcher[value]
                            dict_test_data[test_fields[index]] = value
                            
                        validator = TestValidator(**dict_test_data)
                        validator.is_valid_fields([
                            'status', 'type', 'result',
                        ])
                        
                        validator.extra_validate_to_create_test()

                        list_to_create_test = [key for key in test_fields]
                        list_to_create_test = set(list_to_create_test) - {'user_code'}
                        list_to_create_test = list(list_to_create_test) + ['user']

                        dict_to_create_test = validator.get_data(list_to_create_test)

                        test = Test(**dict_to_create_test)

                        test.code = self.custom_test_code_generator(test.user.code)
                        while (validator.is_code_exist(test.code)):
                            test.code = self.custom_test_code_generator(test.user.code)
                        test.created_by = request.user
                        test.save()

                        # Update user
                        user = test.user
                        if hasattr(user, 'member_x_custom_user'):
                            this_member = user.member_x_custom_user
                            this_member.last_tested = test.created_at

                            if test.result != TestResult.NONE:
                                old_positive_test_now = this_member.positive_test_now
                                new_positive_test_now = self.calculate_new_conclude_from_test(test, old_positive_test_now)
                                this_member.positive_test_now = new_positive_test_now

                                if old_positive_test_now != True and new_positive_test_now == True:
                                    this_member.label = MemberLabel.F0

                                    if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                                        if this_member.quarantined_finish_expected_at == None:
                                            quarantine_ward = this_member.custom_user.quarantine_ward
                                            if quarantine_ward.pandemic:
                                                if this_member.number_of_vaccine_doses < 2:
                                                    remain_qt = quarantine_ward.pandemic.remain_qt_pos_not_vac
                                                else:
                                                    remain_qt = quarantine_ward.pandemic.remain_qt_pos_vac
                                            else:
                                                if this_member.number_of_vaccine_doses < 2:
                                                    remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                                                else:
                                                    remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                                            this_member.quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                                        
                                        # affect other member in this room
                                        this_room = this_member.quarantine_room
                                        if this_room:
                                            members_in_this_room = this_room.member_x_quarantine_room.all()
                                            for member in list(members_in_this_room):
                                                if member.label != MemberLabel.F0:
                                                    member.label = MemberLabel.F1
                                                    member.quarantined_finish_expected_at = None
                                                    member.save()

                                elif old_positive_test_now == True and new_positive_test_now == False:
                                    this_member.positive_tested_before = True

                                    if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                                        # affect other member in this room
                                        this_room = this_member.quarantine_room
                                        if this_room:
                                            other_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=this_member.id)
                                            number_of_other_positive_member_in_this_room = other_members_in_this_room.filter(positive_test_now=True).count()
                                            if number_of_other_positive_member_in_this_room == 0:
                                                quarantine_ward = this_room.quarantine_floor.quarantine_building.quarantine_ward
                                                for member in list(other_members_in_this_room):
                                                    if member.label != MemberLabel.F0:
                                                        if quarantine_ward.pandemic:
                                                            if member.number_of_vaccine_doses < 2:
                                                                remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_not_vac
                                                            else:
                                                                remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_vac
                                                        else:
                                                            if member.number_of_vaccine_doses < 2:
                                                                remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_NOT_VAC', 14))
                                                            else:
                                                                remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_VAC', 10))
                                                        old_quarantined_finish_expected_at = member.quarantined_finish_expected_at
                                                        new_quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                                                        if not old_quarantined_finish_expected_at or old_quarantined_finish_expected_at < new_quarantined_finish_expected_at:
                                                            member.quarantined_finish_expected_at = new_quarantined_finish_expected_at
                                                            member.save()

                                this_member.last_tested_had_result = test.created_at
                            this_member.save()
                        if hasattr(user, 'manager_x_custom_user'):
                            this_manager = user.manager_x_custom_user
                            this_manager.last_tested = test.created_at
                            if test.result != TestResult.NONE:
                                new_positive_test_now = self.calculate_new_conclude_from_test(test, this_manager.positive_test_now)
                                this_manager.positive_test_now = new_positive_test_now
                                this_manager.last_tested_had_result = test.created_at
                            this_manager.save()
                        if hasattr(user, 'staff_x_custom_user'):
                            this_staff = user.staff_x_custom_user
                            this_staff.last_tested = test.created_at
                            if test.result != TestResult.NONE:
                                new_positive_test_now = self.calculate_new_conclude_from_test(test, this_staff.positive_test_now)
                                this_staff.positive_test_now = new_positive_test_now
                                this_staff.last_tested_had_result = test.created_at
                            this_staff.save()

                        if test.result != TestResult.NONE:
                            if test.result == TestResult.POSITIVE:
                                list_user_positive += [test.user]
                            else:
                                list_user_negative += [test.user]

                        test_ids += [test.id]
                    except Exception as e:
                        error[str(row_index + 1)] = str(e)
                        pass

                # Send notification to users
                vntz = pytz.timezone('Asia/Saigon')
                created_at = test.created_at.astimezone(vntz)
                time_string = f'{created_at.hour} giờ {created_at.minute} phút {created_at.second} giây, ngày {created_at.day} tháng {created_at.month} năm {created_at.year}'
                title = 'Kết quả xét nghiệm'
                description_positive = 'Phiếu xét nghiệm lúc ' + time_string + ' có kết quả ' + 'DƯƠNG TÍNH'
                description_negative = 'Phiếu xét nghiệm lúc ' + time_string + ' có kết quả ' + 'ÂM TÍNH'
                create_and_send_noti_to_list_user(title=title, description=description_positive, receive_user_list=list_user_positive, created_by=None)
                create_and_send_noti_to_list_user(title=title, description=description_negative, receive_user_list=list_user_negative, created_by=None)
                
                test_data = Test.objects.filter(id__in=test_ids)
                test_serializer = FilterTestSerializer(test_data, many=True)
                response_data = dict()
                response_data["test_success"] = test_serializer.data
                response_data["test_fail"] = error
                
            else:
                raise exceptions.InvalidArgumentException({'main': messages.FILE_IMPORT_EMPTY})

            return self.response_handler.handle(data=response_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_test(self, request):
        """Get a test

        Args:
            + code: String
        """

        accept_fields = [
            'code',
        ]

        require_fields = [
            'code',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = TestValidator(**accepted_fields)

            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_get_test()

            test = validator.get_field('test')

            test = Test.objects.all().select_related(
                'user__member_x_custom_user', 'user__manager_x_custom_user', 'user__staff_x_custom_user',
                'created_by__member_x_custom_user', 'created_by__manager_x_custom_user', 'created_by__staff_x_custom_user',
                'updated_by__member_x_custom_user', 'updated_by__manager_x_custom_user', 'updated_by__staff_x_custom_user',
            ).get(id=test.id)

            serializer = TestSerializer(test, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_test(self, request):
        """Update a test

        Args:
            + code: String
            - type: String ['QUICK', 'RT-PCR']
            - result: String ['NONE', 'NEGATIVE', 'POSITIVE']
        """

        accept_fields = [
            'code', 'type',
            'result',
        ]

        require_fields = [
            'code',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = TestValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields([
                'type', 'result',
            ])
            
            validator.extra_validate_to_update_test()

            test = validator.get_field('test')

            list_to_update_test = [key for key in accepted_fields.keys()]
            list_to_update_test = set(list_to_update_test) - {'code'}
            list_to_update_test = list(list_to_update_test) + ['status']

            dict_to_update_test = validator.get_data(list_to_update_test)

            for attr, value in dict_to_update_test.items(): 
                setattr(test, attr, value)

            test.updated_by = request.user
            test.save()

            # Update user
            if test.result != TestResult.NONE:
                # if this test has result
                last_test_has_result = Test.objects.filter(user = test.user).filter(~Q(result=TestResult.NONE)).order_by('-created_at')[0]
                if last_test_has_result == test:
                    # This test is the last test of this user that has result
                    if hasattr(test.user, 'member_x_custom_user'):
                        this_member = test.user.member_x_custom_user
                        old_positive_test_now = this_member.positive_test_now
                        new_positive_test_now = self.calculate_new_conclude_from_test(test, this_member.positive_test_now)
                        this_member.positive_test_now = new_positive_test_now

                        if old_positive_test_now != True and new_positive_test_now == True:
                            this_member.label = MemberLabel.F0
                            if not this_member.first_positive_test_date:
                                this_member.first_positive_test_date = test.created_at

                            if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                                if this_member.quarantined_finish_expected_at == None:
                                    quarantine_ward = this_member.custom_user.quarantine_ward
                                    if quarantine_ward.pandemic:
                                        if this_member.number_of_vaccine_doses < 2:
                                            remain_qt = quarantine_ward.pandemic.remain_qt_pos_not_vac
                                        else:
                                            remain_qt = quarantine_ward.pandemic.remain_qt_pos_vac
                                    else:
                                        if this_member.number_of_vaccine_doses < 2:
                                            remain_qt = int(os.environ.get('REMAIN_QT_POS_NOT_VAC', 14))
                                        else:
                                            remain_qt = int(os.environ.get('REMAIN_QT_POS_VAC', 10))
                                    this_member.quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                                
                                # affect other members in this room
                                this_room = this_member.quarantine_room
                                if this_room:
                                    other_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=this_member.id)
                                    for member in list(other_members_in_this_room):
                                        if member.label != MemberLabel.F0:
                                            member.label = MemberLabel.F1
                                            member.quarantined_finish_expected_at = None
                                            member.save()
                        
                        elif old_positive_test_now == True and new_positive_test_now == False:
                            this_member.positive_tested_before = True

                            if this_member.custom_user.status == CustomUserStatus.AVAILABLE:
                                # affect other members in this room
                                this_room = this_member.quarantine_room
                                if this_room:
                                    other_members_in_this_room = this_room.member_x_quarantine_room.all().exclude(id=this_member.id)
                                    number_of_other_positive_member_in_this_room = other_members_in_this_room.filter(positive_test_now=True).count()
                                    if number_of_other_positive_member_in_this_room == 0:
                                        quarantine_ward = this_room.quarantine_floor.quarantine_building.quarantine_ward
                                        for member in list(other_members_in_this_room):
                                            if member.label != MemberLabel.F0:
                                                if quarantine_ward.pandemic:
                                                    if member.number_of_vaccine_doses < 2:
                                                        remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_not_vac
                                                    else:
                                                        remain_qt = quarantine_ward.pandemic.remain_qt_cc_pos_vac
                                                else:
                                                    if member.number_of_vaccine_doses < 2:
                                                        remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_NOT_VAC', 14))
                                                    else:
                                                        remain_qt = int(os.environ.get('REMAIN_QT_CC_POS_VAC', 10))
                                                old_quarantined_finish_expected_at = member.quarantined_finish_expected_at
                                                new_quarantined_finish_expected_at = timezone.now() + datetime.timedelta(days=remain_qt)
                                                if not old_quarantined_finish_expected_at or old_quarantined_finish_expected_at < new_quarantined_finish_expected_at:
                                                    member.quarantined_finish_expected_at = new_quarantined_finish_expected_at
                                                    member.save()

                        this_member.last_tested_had_result = test.created_at
                        this_member.save()

                    if hasattr(test.user, 'manager_x_custom_user'):
                        this_manager = test.user.manager_x_custom_user
                        new_positive_test_now = self.calculate_new_conclude_from_test(test, this_manager.positive_test_now)
                        this_manager.positive_test_now = new_positive_test_now
                        this_manager.last_tested_had_result = test.created_at
                        this_manager.save()
                        
                    if hasattr(test.user, 'staff_x_custom_user'):
                        this_staff = test.user.staff_x_custom_user
                        new_positive_test_now = self.calculate_new_conclude_from_test(test, this_staff.positive_test_now)
                        this_staff.positive_test_now = new_positive_test_now
                        this_staff.last_tested_had_result = test.created_at
                        this_staff.save()

            if test.result != TestResult.NONE:
                vntz = pytz.timezone('Asia/Saigon')
                created_at = test.created_at.astimezone(vntz)
                time_string = f'{created_at.hour} giờ {created_at.minute} phút {created_at.second} giây, ngày {created_at.day} tháng {created_at.month} năm {created_at.year}'
                result_string = 'ÂM TÍNH' if test.result == TestResult.NEGATIVE else 'DƯƠNG TÍNH'
                title = 'Kết quả xét nghiệm'
                description = 'Phiếu xét nghiệm lúc ' + time_string + ' có kết quả ' + result_string
                create_and_send_noti_to_list_user(title=title, description=description, receive_user_list=[test.user], created_by=None)

            serializer = TestSerializer(test, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_test(self, request):
        """Get a list of tests

        Args:
            - user_code: String
            - status: String ['WAITING', 'DONE']
            - result: String ['NONE', 'NEGATIVE', 'POSITIVE']
            - type: String ['QUICK', 'RT-PCR']
            - created_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - created_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - updated_at_max: String vd:'2000-01-26T01:23:45.123456Z'
            - updated_at_min: String vd:'2000-01-26T01:23:45.123456Z'
            - quarantine_ward_id: String
            - quarantine_building_id: String
            - quarantine_floor_id: String
            - quarantine_room_id: String
            - page: int
            - page_size: int
            - search: String
        """

        accept_fields = [
            'user_code', 'status',
            'result', 'type',
            'created_at_max', 'created_at_min',
            'updated_at_max', 'updated_at_min',
            'quarantine_ward_id', 'quarantine_building_id',
            'quarantine_floor_id', 'quarantine_room_id',
            'page', 'page_size', 'search',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = TestValidator(**accepted_fields)

            validator.is_valid_fields([
                'status', 'result', 'type',
                'created_at_max', 'created_at_min',
                'updated_at_max', 'updated_at_min',
            ])
            validator.extra_validate_to_filter_test()

            query_set = Test.objects.all()

            list_to_filter_test = [key for key in accepted_fields.keys()]
            list_to_filter_test = set(list_to_filter_test) - \
            {'page', 'page_size'}

            dict_to_filter_test = validator.get_data(list_to_filter_test)

            # Check ward of sender
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                if hasattr(validator, '_quarantine_ward'):
                    # Sender want filter with ward, building, floor or room
                    if validator.get_field('quarantine_ward') != request.user.quarantine_ward:
                        raise exceptions.AuthenticationException({'quarantine_ward_id': messages.NO_PERMISSION})
                else:
                    dict_to_filter_test['quarantine_ward_id'] = request.user.quarantine_ward.id

            dict_to_filter_test.setdefault('order_by', '-created_at')

            filter = TestFilter(dict_to_filter_test, queryset=query_set)

            query_set = filter.qs

            query_set = query_set.select_related('user__member_x_custom_user', 'user__manager_x_custom_user', 'user__staff_x_custom_user')

            serializer = FilterTestSerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class VaccineAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_vaccine(self, request):
        """Get a vaccine

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineValidator(**accepted_fields)

            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_get_vaccine()

            vaccine = validator.get_field('vaccine')

            serializer = VaccineSerializer(vaccine, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_vaccine(self, request):
        """Create a vaccine

        Args:
            + name: String
            + manufacturer: String
        """

        accept_fields = [
            'name', 'manufacturer',
        ]

        require_fields = [
            'name', 'manufacturer',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_create_vaccine()

            list_to_create_test = [key for key in accepted_fields.keys()]

            dict_to_create_test = validator.get_data(list_to_create_test)

            vaccine = Vaccine(**dict_to_create_test)
            vaccine.save()

            serializer = VaccineSerializer(vaccine, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_vaccine(self, request):
        """Update a vaccine

        Args:
            + id: int
            - name: String
            - manufacturer: String
        """

        accept_fields = [
            'id',
            'name', 'manufacturer',
        ]

        require_fields = [
            'id',
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_update_vaccine()

            vaccine = validator.get_field('vaccine')

            list_to_update_vaccine = [key for key in accepted_fields.keys()]
            list_to_update_vaccine = set(list_to_update_vaccine) - {'id'}

            dict_to_update_vaccine = validator.get_data(list_to_update_vaccine)

            for attr, value in dict_to_update_vaccine.items(): 
                setattr(vaccine, attr, value)

            vaccine.save()

            serializer = VaccineSerializer(vaccine, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_vaccine(self, request):
        """Get a list of vaccine

        Args:
            None
        """

        try:
            query_set = Vaccine.objects.all().order_by('id')

            serializer = VaccineSerializer(query_set, many=True)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

class VaccineDoseAPI(AbstractView):

    permission_classes = [permissions.IsAuthenticated]

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='get', detail=False)
    def get_vaccine_dose(self, request):
        """Get a vaccine dose

        Args:
            + id: int
        """

        accept_fields = [
            'id',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineDoseValidator(**accepted_fields)

            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_get_vaccine_dose()

            vaccine_dose = validator.get_field('vaccine_dose')

            serializer = VaccineDoseSerializer(vaccine_dose, many=False)

            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='create', detail=False)
    def create_vaccine_dose(self, request):
        """Create a vaccine dose

        Args:
            + vaccine_id: int
            + custom_user_code: String
            + injection_date: String vd:'2000-01-26T01:23:45.123456Z'
            - injection_place: String
            - batch_number: String
            - symptom_after_injected: String
        """

        accept_fields = [
            'vaccine_id', 'custom_user_code',
            'injection_date', 'injection_place',
            'batch_number', 'symptom_after_injected',
        ]

        require_fields = [
            'vaccine_id', 'custom_user_code',
            'injection_date',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineDoseValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['injection_date'])
            
            validator.extra_validate_to_create_vaccine_dose()

            custom_user = validator.get_field('custom_user')
            if request.user.role.name == 'MEMBER' and request.user != custom_user:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            list_to_create_vaccine_dose = [key for key in accepted_fields.keys()]

            list_to_create_vaccine_dose = set(list_to_create_vaccine_dose) - \
            {'vaccine_id', 'custom_user_code',}
            list_to_create_vaccine_dose = list(list_to_create_vaccine_dose) + \
            ['vaccine', 'custom_user',]

            dict_to_create_vaccine_dose = validator.get_data(list_to_create_vaccine_dose)

            vaccine_dose = VaccineDose(**dict_to_create_vaccine_dose)
            vaccine_dose.save()

            # Update member/manager/staff.number_of_vaccine_doses

            if hasattr(custom_user, 'member_x_custom_user'):
                member = custom_user.member_x_custom_user
                member.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                member.save()

            if hasattr(custom_user, 'manager_x_custom_user'):
                manager = custom_user.manager_x_custom_user
                manager.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                manager.save()
            
            if hasattr(custom_user, 'staff_x_custom_user'):
                staff = custom_user.staff_x_custom_user
                staff.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                staff.save()

            serializer = VaccineDoseSerializer(vaccine_dose, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='update', detail=False)
    def update_vaccine_dose(self, request):
        """Update a vaccine dose

        Args:
            + id: int
            - vaccine_id: int
            - injection_date: String vd:'2000-01-26T01:23:45.123456Z'
            - injection_place: String
            - batch_number: String
            - symptom_after_injected: String
        """

        accept_fields = [
            'id', 'vaccine_id',
            'injection_date', 'injection_place',
            'batch_number', 'symptom_after_injected',
        ]

        require_fields = [
            'id',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineDoseValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            validator.is_valid_fields(['injection_date'])
            
            validator.extra_validate_to_update_vaccine_dose()

            vaccine_dose = validator.get_field('vaccine_dose')

            list_to_update_vaccine_dose = [key for key in accepted_fields.keys()]
            list_to_update_vaccine_dose = set(list_to_update_vaccine_dose) - \
            {'id', 'vaccine_id'}
            list_to_update_vaccine_dose = list(list_to_update_vaccine_dose) + \
            ['vaccine']

            dict_to_update_vaccine_dose = validator.get_data(list_to_update_vaccine_dose)

            for attr, value in dict_to_update_vaccine_dose.items(): 
                setattr(vaccine_dose, attr, value)

            vaccine_dose.save()

            serializer = VaccineDoseSerializer(vaccine_dose, many=False)
            
            return self.response_handler.handle(data=serializer.data)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='delete', detail=False)
    def delete_vaccine_dose(self, request):
        """Delete a vaccine dose

        Args:
            + id: int
        """

        accept_fields = [
            'id'
        ]

        require_fields = [
            'id'
        ]

        try:
            if request.user.role.name not in ['ADMINISTRATOR', 'SUPER_MANAGER', 'MANAGER', 'STAFF']:
                raise exceptions.AuthenticationException({'main': messages.NO_PERMISSION})

            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineDoseValidator(**accepted_fields)
            validator.is_missing_fields(require_fields)
            
            validator.extra_validate_to_delete_vaccine_dose()

            vaccine_dose = validator.get_field('vaccine_dose')

            custom_user = vaccine_dose.custom_user

            vaccine_dose.delete()

            # Update member/manager/staff.number_of_vaccine_doses

            if hasattr(custom_user, 'member_x_custom_user'):
                member = custom_user.member_x_custom_user
                member.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                member.save()

            if hasattr(custom_user, 'manager_x_custom_user'):
                manager = custom_user.manager_x_custom_user
                manager.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                manager.save()
            
            if hasattr(custom_user, 'staff_x_custom_user'):
                staff = custom_user.staff_x_custom_user
                staff.number_of_vaccine_doses = VaccineDose.objects.filter(custom_user=custom_user).count()
                staff.save()
            
            return self.response_handler.handle(data=messages.SUCCESS)
        except Exception as exception:
            return self.exception_handler.handle(exception)

    @csrf_exempt
    @query_debugger
    @action(methods=['POST'], url_path='filter', detail=False)
    def filter_vaccine_dose(self, request):
        """Get a list of vaccine_doses

        Args:
            - custom_user_code: String
            - injection_date_max: String vd:'2000-01-26T01:23:45.123456Z'
            - injection_date_min: String vd:'2000-01-26T01:23:45.123456Z'
            - page: int
            - page_size: int
            - search: String
        """

        accept_fields = [
            'custom_user_code',
            'injection_date_max', 'injection_date_min',
            'page', 'page_size', 'search',
        ]

        try:
            request_extractor = self.request_handler.handle(request)
            receive_fields = request_extractor.data
            accepted_fields = dict()

            for key in receive_fields.keys():
                if key in accept_fields:
                    accepted_fields[key] = receive_fields[key]

            validator = VaccineDoseValidator(**accepted_fields)

            validator.is_valid_fields([
                'injection_date_max', 'injection_date_min'
            ])
            validator.extra_validate_to_filter_vaccine_dose()

            query_set = VaccineDose.objects.all()

            list_to_filter_vaccine_dose = [key for key in accepted_fields.keys()]
            list_to_filter_vaccine_dose = set(list_to_filter_vaccine_dose) - \
            {'page', 'page_size'}

            dict_to_filter_vaccine_dose = validator.get_data(list_to_filter_vaccine_dose)

            dict_to_filter_vaccine_dose.setdefault('order_by', 'injection_date')

            filter = VaccineDoseFilter(dict_to_filter_vaccine_dose, queryset=query_set)

            query_set = filter.qs

            query_set = query_set.select_related()

            serializer = FilterVaccineDoseSerializer(query_set, many=True)

            paginated_data = paginate_data(request, serializer.data)

            return self.response_handler.handle(data=paginated_data)
        except Exception as exception:
            return self.exception_handler.handle(exception)
